"""Test script para diagnosticar problema de openpyxl"""
import sys
print(f"Python: {sys.version}")
print(f"Python executable: {sys.executable}")
print(f"sys.path:")
for p in sys.path:
    print(f"  {p}")

print("\n--- Intentando importar openpyxl ---")
try:
    import openpyxl
    print(f"✓ openpyxl importado OK")
    print(f"  Version: {openpyxl.__version__}")
    print(f"  Location: {openpyxl.__file__}")
except Exception as e:
    print(f"✗ Error importando openpyxl: {e}")
    sys.exit(1)

print("\n--- Intentando importar pandas ---")
try:
    import pandas as pd
    print(f"✓ pandas importado OK")
    print(f"  Version: {pd.__version__}")
    print(f"  Location: {pd.__file__}")
except Exception as e:
    print(f"✗ Error importando pandas: {e}")
    sys.exit(1)

print("\n--- Verificando pandas.io.excel ---")
try:
    import pandas.io.excel
    print(f"✓ pandas.io.excel OK")
    print(f"  _OpenpyxlReader disponible: {hasattr(pd.io.excel, '_OpenpyxlReader')}")
except Exception as e:
    print(f"✗ Error: {e}")

print("\n--- Intentando leer Excel con openpyxl ---")
try:
    # Crear un Excel de prueba
    import io
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'Nombre'
    ws['B1'] = 'Valor'
    ws['A2'] = 'Test'
    ws['B2'] = 123
    
    # Guardar a BytesIO
    excel_bytes = io.BytesIO()
    wb.save(excel_bytes)
    excel_bytes.seek(0)
    
    # Intentar leer con pandas
    df = pd.read_excel(excel_bytes, engine='openpyxl')
    print(f"✓ pandas.read_excel con engine='openpyxl' funciona OK")
    print(f"  DataFrame shape: {df.shape}")
    print(f"  Datos:\n{df}")
except Exception as e:
    print(f"✗ Error leyendo Excel con pandas: {e}")
    import traceback
    traceback.print_exc()
    sys.exit(1)

print("\n=== TODO OK ===")
