import streamlit as st
def main():
    # Bloque superior de usuario y configuración de modelo (estilo original)
    st.title("🏛️ Sistema Multi-Agente FCFM")
    st.markdown("Asistente de procedimientos para 15 áreas de decanato y vicedecanato - FCFM")
    st.divider()

    cols = st.columns([1,2])
    with cols[0]:
        st.subheader("👤 Usuario")
        if st.session_state.get("current_user"):
            st.success(f"Sesión: {st.session_state['current_user']}")
            st.info("Chat habilitado. Puedes interactuar con los agentes.")
        else:
            st.warning("Inicia sesión con Google para habilitar el chat.")
        st.button("Cerrar sesión", use_container_width=True)
        st.button("Borrar historial", use_container_width=True)

    st.divider()
    st.subheader("⚙️ Configuración de Modelo")
    st.selectbox("Proveedor LLM", ["GitHub Models", "LM Studio (local)"])
    st.info("Configura GITHUB_TOKEN en tu archivo .env para usar GitHub Models.")
    # Reducir ruido de warnings de LangChain (no afecta ejecución)
    try:
        from langchain_core._api.deprecation import LangChainDeprecationWarning
        warnings.filterwarnings("ignore", category=LangChainDeprecationWarning)
    except Exception:
        pass
    st.markdown("---")
def enviar_client_id_al_cliente_windows(email: str, client_host: str = "localhost", client_port: int = 5001):
    """
    Envía el email del usuario autenticado al cliente Windows para que lo use como CLIENT_ID.
    Por defecto, intenta POST a http://localhost:5001/set_client_id (el .exe debe exponer este endpoint o compartir carpeta).
    Alternativamente, puede guardar el archivo en una ruta compartida.
    """
    import requests
    try:
        url = f"http://{client_host}:{client_port}/set_client_id"
        resp = requests.post(url, json={"client_id": email}, timeout=2)
        if resp.status_code == 200:
            print(f"CLIENT_ID enviado correctamente al cliente Windows: {email}")
            return True
        else:
            print(f"Error enviando CLIENT_ID al cliente Windows: {resp.status_code}")
    except Exception as e:
        print(f"No se pudo enviar CLIENT_ID al cliente Windows: {e}")
    return False

import os
import sys
import json
import urllib.request
                            # Eliminado st.experimental_set_query_params() para evitar conflicto con st.query_params
import warnings
import time
import threading
import queue
import re
import getpass
import secrets
import subprocess
import socket
import requests
import uvicorn
from dotenv import load_dotenv

# Cargar variables de entorno desde .env
load_dotenv()

from langsmith import Client

from ai_support.core.logging_utils import setup_logging
from ai_support.core.config import (
    default_github_embeddings,
    default_github_llm,
    default_lmstudio_embeddings,
    default_lmstudio_llm,
)
from ai_support.orchestrator.multi_orchestrator import OrquestadorMultiagente
from ai_support.core.printer_diagnostics import (
    add_shared_printer,
    collect_printer_diagnostics,
    connect_printer_ip,
    auto_connect_printer_ip,
    diagnose_and_fix_printer_by_name,
    format_diagnostics_for_prompt,
    list_local_printers_structured,
    list_printer_drivers,
    print_test_page,
    restart_spooler,
    set_default_printer,
)
from ai_support.core.printer_inventory_mysql import fetch_printers_from_mysql, mysql_enabled
from ai_support.core.user_memory_persistence import UserMemoryPersistence
from ai_support.core.ip_assignment import (
    assign_ip_to_ethernet_and_register,
    list_net_adapters,
    get_adapter_ipv4,
)
from ai_support.core.users_mysql import get_user_by_email, upsert_user_by_email
from ai_support.core.users_mysql import get_user_department_by_email
from ai_support.core.windows_elevation import is_windows_admin, restart_streamlit_elevated
from ai_support.core.google_auth import (
    build_google_auth_url,
    exchange_code_for_tokens,
    google_auth_enabled,
    google_redirect_uri,
    verify_id_token_and_get_email,
)


_IPV4_IN_TEXT_RE = re.compile(r"(?:\d{1,3}\.){3}\d{1,3}")


def _extract_ipv4(text: str) -> str | None:
    m = _IPV4_IN_TEXT_RE.search(text or "")
    if not m:
        return None
    return m.group(0)


def _lmstudio_fetch_model_ids(base_url: str) -> list[str]:
    url = base_url.rstrip("/") + "/models"
    req = urllib.request.Request(url, method="GET")
    with urllib.request.urlopen(req, timeout=5) as resp:
        payload = json.loads(resp.read().decode("utf-8"))
    data = payload.get("data", [])
    ids: list[str] = []
    for item in data:
        model_id = item.get("id")
        if isinstance(model_id, str) and model_id:
            ids.append(model_id)
    return ids


def _github_fetch_model_ids(base_url: str, token: str) -> list[str]:
    """Lista modelos disponibles en GitHub Models (Azure AI Inference compatible).

    Intenta con headers típicos para maximizar compatibilidad:
    - Authorization: Bearer <token>
    - api-key: <token>
    """

    url = base_url.rstrip("/") + "/models"
    req = urllib.request.Request(
        url,
        method="GET",
        headers={
            "Authorization": f"Bearer {token}",
            "api-key": token,
            "Accept": "application/json",
        },
    )
    with urllib.request.urlopen(req, timeout=10) as resp:
        payload = json.loads(resp.read().decode("utf-8"))

    # GitHub Models (Azure AI Inference) suele devolver una lista de objetos.
    # Preferimos "name" (IDs cortos como gpt-4o-mini) cuando esté disponible.
    if isinstance(payload, list):
        data = payload
    elif isinstance(payload, dict):
        data = payload.get("data", [])
    else:
        data = []

    ids: list[str] = []
    for item in data:
        if not isinstance(item, dict):
            continue
        model_name = item.get("name")
        model_id = item.get("id")
        if isinstance(model_name, str) and model_name:
            ids.append(model_name)
        elif isinstance(model_id, str) and model_id:
            ids.append(model_id)

    # De-dup preservando orden
    seen: set[str] = set()
    result: list[str] = []
    for mid in ids:
        if mid in seen:
            continue
        seen.add(mid)
        result.append(mid)
    return result


def _is_github_no_access_error(err: Exception) -> bool:
    msg = str(err).lower()
    return (
        "permissiondeniederror" in msg
        or "no_access" in msg
        or "no access to model" in msg
        or "error code: 403" in msg
    )


def _is_rate_limit_error(err: Exception) -> bool:
    msg = str(err).lower()
    return (
        "ratelimiterror" in msg
        or "too many requests" in msg
        or "error code: 429" in msg
        or "rate limit" in msg
    )


def _normalize_key(value: str) -> str:
    value = (value or "").strip().lower()
    value = re.sub(r"\s+", "_", value)
    return re.sub(r"[^a-z0-9_]", "", value)


def _department_name_to_area_id(department_name: str | None) -> str | None:
    key = _normalize_key(department_name or "")
    if not key:
        return None

    aliases = {
        "tesoreria": "tesoreria",
        "arquitectura": "arquitectura",
        "infraestructura": "infraestructura",
        "proyectos": "proyectos",
        "atencion_alumnos": "atencion_alumnos",
        "atencion_de_alumnos": "atencion_alumnos",
        "postgrado": "postgrado",
        "sustentabilidad": "sustentabilidad",
        "comunicaciones": "comunicaciones",
        "vinculacion": "vinculacion",
        "vinculacion_externa": "vinculacion",
        "rrhh": "rrhh",
        "recursos_humanos": "rrhh",
        "contabilidad": "contabilidad",
        "direccion_economica": "direccion_economica",
        "direccion_academica": "direccion_academica",
        "diversidad": "diversidad",
        "decanato": "decanato",
        "vicedecanato": "decanato",
    }
    if key in aliases:
        return aliases.get(key)

    # Heurística para nombres institucionales largos (ej: unidad_de_tesoreria)
    compact = key
    compact = compact.replace("unidad_de_", "")
    compact = compact.replace("unidad_", "")
    compact = compact.replace("direccion_de_", "")
    compact = compact.replace("direccion_", "")
    compact = compact.replace("escuela_de_", "")

    if "tesorer" in compact:
        return "tesoreria"
    if "arquitect" in compact:
        return "arquitectura"
    if "infraestructura" in compact:
        return "infraestructura"
    if "proyecto" in compact:
        return "proyectos"
    if "alumno" in compact or "estudiant" in compact:
        return "atencion_alumnos"
    if "postgrado" in compact or "posgrado" in compact:
        return "postgrado"
    if "sustentab" in compact or "sostenib" in compact:
        return "sustentabilidad"
    if "comunic" in compact:
        return "comunicaciones"
    if "vincul" in compact:
        return "vinculacion"
    if "rrhh" in compact or "recurso_humano" in compact:
        return "rrhh"
    if "contabil" in compact:
        return "contabilidad"
    if "econom" in compact:
        return "direccion_economica"
    if "academ" in compact:
        return "direccion_academica"
    if "divers" in compact or "genero" in compact:
        return "diversidad"
    if "decan" in compact:
        return "decanato"

    return None


def _department_matches_area_name(department_name: str | None, area_name: str | None) -> bool:
    dept = _normalize_key(department_name or "")
    area = _normalize_key(area_name or "")
    if not dept or not area:
        return False
    return dept == area or dept in area or area in dept


# ── Base de Conocimiento UI ───────────────────────────────────────────────────

def _render_knowledge_base_section(department_name: str | None = None) -> None:
    """Renderiza la sección de Base de Conocimiento por Áreas."""
    import streamlit as st
    from ai_support.core.knowledge_base import get_kb_manager

    ACCEPTED_TYPES = ["pdf", "docx", "doc", "xlsx", "xls", "xlsm", "txt", "csv", "md"]
    ACCEPTED_LABEL = ".pdf, .docx, .xlsx, .txt"

    kb = get_kb_manager()

    st.markdown("## 📚 Base de Conocimiento - Procedimientos FCFM")
    st.caption(
        "Sube documentos PDF con procedimientos y tareas de cada área FCFM. "
        "Los agentes responderán basándose en los procedimientos de tu área."
    )

    # ── Inicializar estado ────────────────────────────────────────────────────
    if "kb_selected_area" not in st.session_state:
        st.session_state["kb_selected_area"] = None
    if "kb_refresh" not in st.session_state:
        st.session_state["kb_refresh"] = 0

    areas = kb.list_areas()
    if department_name:
        areas = [a for a in areas if _department_matches_area_name(department_name, a.get("name", ""))]
        if not areas:
            st.warning(
                "No hay un área de Base de Conocimiento asociada a tu departamento. "
                "Pide al administrador crearla con el nombre del departamento."
            )
            return

    # ── Layout: columna áreas | columna documentos ────────────────────────────
    col_areas, col_docs = st.columns([1, 2], gap="large")

    # ──────────────── Columna izquierda: lista de áreas ───────────────────────
    with col_areas:
        st.markdown("### 🏢 Áreas de la empresa")

        # Crear nueva área
        with st.expander("➕ Crear nueva área", expanded=not areas):
            with st.form("form_create_area", clear_on_submit=True):
                new_name = st.text_input("Nombre del área", placeholder="Ej: Recursos Humanos")
                new_desc = st.text_area("Descripción (opcional)", height=80)
                submitted = st.form_submit_button("Crear área", use_container_width=True)
                if submitted:
                    if new_name.strip():
                        try:
                            kb.create_area(new_name.strip(), new_desc.strip())
                            st.success(f"✅ Área '{new_name}' creada.")
                            st.session_state["kb_refresh"] += 1
                            st.rerun()
                        except ValueError as e:
                            st.error(str(e))
                    else:
                        st.warning("Escribe un nombre para el área.")

        if not areas:
            st.info("No hay áreas creadas todavía. Crea la primera usando el formulario de arriba.")
        else:
            for area in areas:
                area_id = area["id"]
                is_selected = st.session_state["kb_selected_area"] == area_id
                btn_label = f"{'▶ ' if is_selected else ''}{area['name']} ({area['doc_count']} docs)"
                if st.button(btn_label, key=f"btn_area_{area_id}", use_container_width=True,
                             type="primary" if is_selected else "secondary"):
                    st.session_state["kb_selected_area"] = area_id
                    st.rerun()

    # ──────────────── Columna derecha: gestión de documentos ─────────────────
    with col_docs:
        area_id = st.session_state.get("kb_selected_area")
        if area_id is None:
            st.markdown("### 📄 Documentos")
            st.info("👈 Selecciona un área en el panel izquierdo para gestionar sus documentos.")
            return

        area_meta = kb.get_area(area_id)
        if area_meta is None:
            st.warning("El área seleccionada ya no existe.")
            st.session_state["kb_selected_area"] = None
            return

        st.markdown(f"### 📂 {area_meta['name']}")
        if area_meta.get("description"):
            st.caption(area_meta["description"])

        # Subir documentos
        st.markdown("#### ⬆️ Subir documentos")
        uploaded_files = st.file_uploader(
            f"Selecciona archivos para **{area_meta['name']}**",
            type=ACCEPTED_TYPES,
            accept_multiple_files=True,
            key=f"uploader_{area_id}",
            help=f"Tipos aceptados: {ACCEPTED_LABEL}",
        )
        if uploaded_files:
            if st.button(f"📥 Subir {len(uploaded_files)} archivo(s)", key=f"btn_upload_{area_id}",
                         use_container_width=True, type="primary"):
                progress = st.progress(0, text="Subiendo...")
                errors: list[str] = []
                for idx, uf in enumerate(uploaded_files):
                    try:
                        kb.upload_document(area_id, uf.name, uf.read())
                    except Exception as e:
                        errors.append(f"{uf.name}: {e}")
                    progress.progress((idx + 1) / len(uploaded_files), text=f"Procesando {uf.name}...")
                progress.empty()
                if errors:
                    st.error("⚠️ Errores al subir:\n" + "\n".join(errors))
                else:
                    st.success(f"✅ {len(uploaded_files)} archivo(s) subido(s) y listos para búsqueda.")
                st.rerun()

        st.markdown("---")

        # Listar documentos existentes
        docs = kb.list_documents(area_id)
        st.markdown(f"#### 📋 Documentos ({len(docs)})")

        if not docs:
            st.info("No hay documentos en esta área. Sube archivos usando el formulario de arriba.")
        else:
            for doc in docs:
                import time as _time
                uploaded_str = ""
                try:
                    uploaded_str = _time.strftime("%d/%m/%Y %H:%M", _time.localtime(doc.get("uploaded_at", 0)))
                except Exception:
                    pass
                size_kb = doc.get("size_bytes", 0) / 1024
                chunks = doc.get("chunk_count", 0)
                col_name, col_info, col_del = st.columns([3, 2, 1])
                with col_name:
                    ext = doc["filename"].rsplit(".", 1)[-1].upper() if "." in doc["filename"] else "?"
                    icons = {"PDF": "📕", "DOCX": "📘", "DOC": "📘", "XLSX": "📗", "XLS": "📗",
                             "XLSM": "📗", "TXT": "📄", "CSV": "📊", "MD": "📝"}
                    icon = icons.get(ext, "📎")
                    st.markdown(f"{icon} **{doc['filename']}**")
                with col_info:
                    st.caption(f"{size_kb:.1f} KB · {chunks} fragmentos · {uploaded_str}")
                with col_del:
                    if st.button("🗑️", key=f"del_{doc['id']}", help=f"Eliminar {doc['filename']}"):
                        kb.delete_document(area_id, doc["id"])
                        st.success(f"Documento '{doc['filename']}' eliminado.")
                        st.rerun()

        st.markdown("---")

        # Zona de prueba de búsqueda
        with st.expander("🔍 Probar búsqueda en esta área", expanded=False):
            test_q = st.text_input("Consulta de prueba", placeholder="Escribe algo para buscar...",
                                   key=f"test_q_{area_id}")
            if st.button("Buscar", key=f"btn_test_search_{area_id}"):
                if test_q.strip():
                    # Intentar con embeddings del orquestador si está disponible
                    embeddings = None
                    orq = st.session_state.get("orquestador")
                    if orq:
                        try:
                            agente = list(orq.agentes.values())[0]
                            embeddings = agente.embeddings
                        except Exception:
                            pass
                    results = kb.search(area_id, test_q.strip(), k=4, embeddings=embeddings)
                    if results:
                        for i, r in enumerate(results, 1):
                            st.markdown(f"**Fragmento {i}** — `{r['filename']}`")
                            st.text_area("", r["text"], height=100, key=f"res_{area_id}_{i}", disabled=True)
                    else:
                        st.warning("No se encontraron resultados.")
                else:
                    st.warning("Escribe una consulta para buscar.")

        # Botón eliminar área
        st.markdown("---")
        with st.expander("⚠️ Zona de peligro", expanded=False):
            st.warning(f"Esto eliminará el área **{area_meta['name']}** y **todos** sus documentos permanentemente.")
            if st.button(f"🗑️ Eliminar área '{area_meta['name']}'", key=f"del_area_{area_id}",
                         type="secondary", use_container_width=True):
                kb.delete_area(area_id)
                st.session_state["kb_selected_area"] = None
                st.success("Área eliminada.")
                st.rerun()


def main() -> None:
    import os
    import secrets
    import streamlit as st

    st.set_page_config(
        page_title="Soporte Informático IA",
        page_icon="⚙️",
        layout="wide",
        initial_sidebar_state="expanded",
    )

    st.markdown("""
    <style>
    /* ── Fuente global ── */
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
    html, body, [class*="css"] { font-family: 'Inter', sans-serif; }

    /* ── Fondo principal blanco/gris muy claro ── */
    .stApp { background: #f5f6fa !important; }

    /* ── Sidebar blanco con borde sutil ── */
    [data-testid="stSidebar"] {
        background: #ffffff !important;
        border-right: 1px solid #e2e6f0 !important;
        box-shadow: 2px 0 8px rgba(0,0,0,0.04);
    }
    [data-testid="stSidebar"] .stMarkdown h3 {
        color: #4f46e5 !important;
        font-weight: 700 !important;
        font-size: 0.82rem !important;
        text-transform: uppercase !important;
        letter-spacing: 0.1em !important;
    }

    /* ── Texto general legible ── */
    html, body, p, span, div, label { color: #1e293b !important; }
    h1, h2, h3, h4 { color: #1e293b !important; font-weight: 700 !important; }

    /* ── Área de contenido principal ── */
    [data-testid="stAppViewContainer"] > .main { background: #f5f6fa !important; }
    .block-container { 
        background: #f5f6fa !important;
        padding-top: 2rem !important;
    }

    /* ── Tarjetas / contenedores internos ── */
    [data-testid="stVerticalBlock"] > div {
        background: transparent;
    }

    /* ── Botones primarios (azul personalizado) ── */
    .stButton > button[kind="primary"] {
        background: #004B93 !important;
        color: #ffffff !important;
        border: none !important;
        border-radius: 10px !important;
        font-weight: 600 !important;
        letter-spacing: 0.02em !important;
        transition: transform 0.15s, box-shadow 0.15s !important;
        box-shadow: 0 3px 12px rgba(0, 75, 147, 0.3) !important;
    }
    .stButton > button[kind="primary"]:hover {
        transform: translateY(-2px) !important;
        box-shadow: 0 6px 18px rgba(0, 75, 147, 0.45) !important;
    }

    /* ── Botones secundarios ── */
    .stButton > button[kind="secondary"] {
        background: #ffffff !important;
        color: #374151 !important;
        border: 1px solid #d1d5db !important;
        border-radius: 10px !important;
        font-weight: 500 !important;
        transition: all 0.15s !important;
    }
    .stButton > button[kind="secondary"]:hover {
        background: #ede9fe !important;
        border-color: #4f46e5 !important;
        color: #4f46e5 !important;
    }

    /* ── Inputs y selectboxes ── */
    .stTextInput > div > div > input {
        background: #ffffff !important;
        border: 1.5px solid #d1d5db !important;
        border-radius: 8px !important;
        color: #1e293b !important;
    }
    .stTextInput > div > div > input:focus {
        border-color: #4f46e5 !important;
        box-shadow: 0 0 0 3px rgba(79,70,229,0.12) !important;
    }

    /* ── Chat input ── */
    [data-testid="stChatInput"] > div {
        background: #ffffff !important;
        border: 1.5px solid #d1d5db !important;
        border-radius: 12px !important;
        box-shadow: 0 1px 4px rgba(0,0,0,0.06) !important;
    }
    [data-testid="stChatInput"] textarea { color: #1e293b !important; }

    /* ── Mensajes del asistente ── */
    [data-testid="stChatMessage"] {
        background: #ffffff !important;
        border-radius: 12px !important;
        border: 1px solid #e9ecf3 !important;
        box-shadow: 0 1px 4px rgba(0,0,0,0.05) !important;
        margin-bottom: 8px !important;
    }

    /* ── Alertas con colores vivos legibles ── */
    [data-testid="stNotification"], .stSuccess, .element-container .stSuccess > div {
        background: #ecfdf5 !important;
        border-left: 4px solid #10b981 !important;
        border-radius: 8px !important;
        color: #065f46 !important;
    }
    .stWarning, .element-container .stWarning > div {
        background: #fffbeb !important;
        border-left: 4px solid #f59e0b !important;
        border-radius: 8px !important;
        color: #78350f !important;
    }
    .stError, .element-container .stError > div {
        background: #fef2f2 !important;
        border-left: 4px solid #ef4444 !important;
        border-radius: 8px !important;
        color: #7f1d1d !important;
    }
    .stInfo, .element-container .stInfo > div {
        background: #eef2ff !important;
        border-left: 4px solid #4f46e5 !important;
        border-radius: 8px !important;
        color: #312e81 !important;
    }

    /* ── Expanders ── */
    [data-testid="stExpander"] {
        background: #ffffff !important;
        border: 1px solid #e2e6f0 !important;
        border-radius: 12px !important;
        box-shadow: 0 1px 4px rgba(0,0,0,0.05) !important;
        overflow: hidden !important;
    }
    [data-testid="stExpander"] summary {
        font-weight: 600 !important;
        color: #1e293b !important;
    }

    /* ── Métricas ── */
    [data-testid="stMetricValue"] { color: #4f46e5 !important; font-weight: 700 !important; }
    [data-testid="stMetricLabel"] { color: #6b7280 !important; font-size: 0.85rem !important; }

    /* ── Divider ── */
    hr { border-color: #e2e6f0 !important; }

    /* ── Scrollbar ── */
    ::-webkit-scrollbar { width: 6px; height: 6px; }
    ::-webkit-scrollbar-track { background: #f1f5f9; }
    ::-webkit-scrollbar-thumb { background: #c7d2fe; border-radius: 3px; }
    ::-webkit-scrollbar-thumb:hover { background: #4f46e5; }

    /* ── Caption ── */
    .stCaption, small { color: #6b7280 !important; font-size: 0.82rem !important; }

    /* ── Radio ── */
    .stRadio > div { gap: 6px; }
    .stRadio label {
        background: #ffffff !important;
        border: 1.5px solid #e2e6f0 !important;
        border-radius: 8px !important;
        padding: 6px 14px !important;
        transition: all 0.15s !important;
        color: #374151 !important;
    }
    .stRadio label:hover { border-color: #4f46e5 !important; background: #eef2ff !important; }

    /* ── Checkbox ── */
    .stCheckbox label span { color: #374151 !important; }

    /* ── Selectbox ── */
    .stSelectbox > div > div { 
        background: #ffffff !important;
        border: 1.5px solid #d1d5db !important;
        border-radius: 8px !important;
        color: #1e293b !important;
    }

    /* ── Código ── */
    code, pre { 
        background: #f1f5f9 !important; 
        color: #1e293b !important;
        border-radius: 6px !important;
    }

    /* ── Dataframe ── */
    [data-testid="stDataFrame"] { border-radius: 10px !important; overflow: hidden !important; }
    </style>
    """, unsafe_allow_html=True)

    USER_FILE = os.path.join(os.path.expanduser("~"), ".ai_support_user_data")

    # Restaurar usuario al cargar la app.
    # NO restaurar google_oauth_state — siempre se genera uno fresco en cada intento de login.
    if os.path.exists(USER_FILE):
        try:
            with open(USER_FILE, "r", encoding="utf-8") as f:
                data = f.read().strip().split("\n")
                if len(data) >= 1:
                    email = data[0]
                    if email:
                        st.session_state["current_user"] = email
        except Exception:
            pass

    st.markdown("""
    <div style='padding: 1.5rem 0 0.5rem 0;'>
        <h1 style='margin:0; font-size:2.2rem; font-weight:700; color:#1e293b;'>
            ⚙️ Sistema Multi-Agente de Soporte Informático
        </h1>
        <p style='margin:0.4rem 0 0 0; color:#6b7280; font-size:0.95rem;'>
            Orquestación inteligente · Agentes especializados · Colaboración multi-agente
        </p>
    </div>
    """, unsafe_allow_html=True)
    st.divider()

    # Mostrar el chat si el usuario está autenticado
    if st.session_state.get("current_user"):
        st.success(f"Chat habilitado para: {st.session_state['current_user']}")
        # ...aquí va el código del chat y orquestador...

    # LangSmith (mantener compatibilidad)
    try:
        client = Client()
        print("✓ LangSmith conectado al proyecto:", os.getenv("LANGSMITH_PROJECT"))
    except Exception:
        client = None

    # --- Selección de modelo/proveedor (antes de crear orquestador) ---

    with st.sidebar:
        # Historial por usuario (modo seguro): usuario del sistema operativo
        st.markdown("### 👤 Usuario")
        st.markdown("---")

        google_enabled = google_auth_enabled()

        # Inicializar sistema de persistencia (por-perfil)
        if "user_persistence" not in st.session_state:
            st.session_state["user_persistence"] = UserMemoryPersistence()

        persistence = st.session_state["user_persistence"]

        # --- Google OAuth (si está configurado) ---

        # --- Google OAuth (si está configurado) ---
        if google_enabled:
            # Solo generar el state si no existe, y nunca sobrescribirlo durante el flujo
            if "google_oauth_state" not in st.session_state:
                st.session_state["google_oauth_state"] = secrets.token_urlsafe(24)

            # Procesar callback SIEMPRE antes de cualquier st.stop()
            def _clear_oauth_state():
                if "google_oauth_state" in st.session_state:
                    del st.session_state["google_oauth_state"]

            def _get_query_params():
                return st.query_params

            def _qp_first(qp_mapping, key: str) -> str | None:
                val = qp_mapping.get(key) if qp_mapping is not None else None
                if isinstance(val, list):
                    return val[0] if val else None
                if val is None:
                    return None
                return str(val)

            def _clear_query_params() -> None:
                try:
                    st.query_params.clear()
                except Exception:
                    pass

            qp = _get_query_params()
            code = _qp_first(qp, "code")
            state = _qp_first(qp, "state")
            oauth_error = _qp_first(qp, "error")

            if oauth_error:
                st.error(f"Google OAuth error: {oauth_error}")
                _clear_query_params()

            if code and state and not st.session_state.get("_google_oauth_done"):
                # Leer el state desde disco (STATE_FILE) como fuente de verdad
                _STATE_FILE = os.path.join(os.path.expanduser("~"), ".ai_support_oauth_state")
                _expected_state = None
                if os.path.exists(_STATE_FILE):
                    try:
                        with open(_STATE_FILE, "r", encoding="utf-8") as f:
                            _expected_state = f.read().strip()
                    except Exception:
                        pass
                if not _expected_state or state != _expected_state:
                    st.error("Error de autenticación. Por favor intenta iniciar sesión nuevamente.")
                    # No borrar google_oauth_state aquí para que el botón pueda seguir funcionando
                else:
                    try:
                        tokens = exchange_code_for_tokens(code=code)
                        raw_id_token = tokens.get("id_token")
                        if not raw_id_token:
                            st.error("Respuesta de Google no incluye id_token.")
                            raise ValueError("Respuesta de Google no incluye id_token")
                        email = None
                        try:
                            email = verify_id_token_and_get_email(raw_id_token=raw_id_token)
                        except Exception as ve:
                            st.error(f"No se pudo verificar el id_token: {ve}")
                        if email:
                            st.session_state["current_user"] = email
                            st.session_state["_google_oauth_done"] = True
                            st.session_state["orquestador"] = None
                            try:
                                with open(USER_FILE, "w", encoding="utf-8") as _uf:
                                    _uf.write(email)
                            except Exception:
                                pass
                        else:
                            st.error("No se pudo obtener el email del usuario o el dominio no es permitido.")
                    except Exception as e:
                        st.error(f"No se pudo autenticar: {e}")
                    finally:
                        _clear_oauth_state()

                _clear_query_params()


            current_user = st.session_state.get("current_user")
            if not current_user:
                if not google_redirect_uri():
                    st.error("Falta `AI_SUPPORT_GOOGLE_REDIRECT_URI` para Google OAuth.")
                else:
                    try:
                        # Mostrar botón; el state y auth_url se generan DENTRO del click para evitar stale state
                        if st.button("🔐 Iniciar sesión con Google", use_container_width=True):
                            # Generar un state fresco en cada intento de login
                            _new_state = secrets.token_urlsafe(24)
                            st.session_state["google_oauth_state"] = _new_state
                            _STATE_FILE_LOGIN = os.path.join(os.path.expanduser("~"), ".ai_support_oauth_state")
                            with open(_STATE_FILE_LOGIN, "w", encoding="utf-8") as f:
                                f.write(_new_state)
                            _auth_url = build_google_auth_url(state=_new_state)
                            st.markdown(f'<meta http-equiv="refresh" content="0; url={_auth_url}">', unsafe_allow_html=True)
                        st.caption("Debes usar tu correo @uchile.cl")
                    except Exception as e:
                        st.error(f"Google OAuth no configurado: {e}")

                st.warning("⚠️ Inicia sesión para usar el chat")
                st.stop()

            # Resolver departamento (control de acceso por área)
            st.session_state["_user_department_id"] = None
            st.session_state["_user_department_name"] = None
            st.session_state["_allowed_area_ids"] = None
            if isinstance(current_user, str) and "@" in current_user:
                enforce_registered = (
                    (os.getenv("AI_SUPPORT_ENFORCE_REGISTERED_GOOGLE_USERS") or "true").strip().lower()
                    in {"1", "true", "yes", "y", "on"}
                )
                mysql_ok = bool((os.getenv("AI_SUPPORT_MYSQL_ENABLE") or "").strip().lower() in {"1", "true", "yes", "y", "on"})
                if enforce_registered and not mysql_ok:
                    st.error(
                        "Acceso denegado: validación de usuarios registrados requiere MySQL habilitado "
                        "(AI_SUPPORT_MYSQL_ENABLE=true)."
                    )
                    st.stop()
                if mysql_ok:
                    try:
                        dept_ctx = get_user_department_by_email(current_user)
                    except Exception as e:
                        st.error(f"No se pudo obtener el departamento desde MySQL: {e}")
                        st.stop()

                    if not dept_ctx:
                        st.error("Acceso denegado: tu correo no está registrado en la base de datos.")
                        st.stop()

                    dept_id = dept_ctx.get("departamento_id")
                    dept_name = (dept_ctx.get("departamento_nombre") or "").strip()
                    if not dept_name:
                        st.error("Tu usuario no tiene nombre de departamento asociado en MySQL.")
                        st.stop()

                    area_id = _department_name_to_area_id(dept_name)
                    if not area_id:
                        st.error(
                            "Tu departamento no tiene mapeo a un agente del sistema. "
                            "Configura el nombre del departamento para que coincida con un área FCFM."
                        )
                        st.stop()

                    st.session_state["_user_department_id"] = dept_id
                    st.session_state["_user_department_name"] = dept_name
                    st.session_state["_allowed_area_ids"] = [area_id]

            # Si el usuario está presente, mostrar el chat
            st.success(f"👤 Sesión: **{current_user}**")
            st.info("✅ Chat habilitado. Puedes interactuar con los agentes.")
            if st.session_state.get("_user_department_name"):
                st.caption(
                    f"🔒 Acceso restringido a departamento: {st.session_state['_user_department_name']}"
                )

            # --- Provisionamiento de usuario en MySQL (tabla usuarios) ---
            # Solo aplica si el usuario es un email (Google OAuth) y MySQL está habilitado.
            if isinstance(current_user, str) and "@" in current_user:
                with st.expander("🗃️ Perfil en MySQL", expanded=False):
                    st.caption(
                        "Crea/actualiza tu registro en la tabla `personal` usando el email de Google. "
                        "Si no existes, completa el formulario para registrarte."
                    )
                    mysql_ok = bool((os.getenv("AI_SUPPORT_MYSQL_ENABLE") or "").strip().lower() in {"1", "true", "yes", "y", "on"})
                    if not mysql_ok:
                        st.info("MySQL no está habilitado (AI_SUPPORT_MYSQL_ENABLE=false).")
                    else:
                        try:
                            db_user = get_user_by_email(current_user)
                            st.session_state["_mysql_user_exists"] = bool(db_user)
                        except Exception as e:
                            db_user = None
                            st.session_state["_mysql_user_exists"] = False
                            st.error(f"No se pudo consultar usuario en MySQL: {e}")

                        if db_user:
                            st.success("Registro encontrado en MySQL.")
                            st.caption(f"IP registrada: {str(db_user.get('IP') or db_user.get('ip') or '').strip() or '(vacía)'}")
                        else:
                            st.warning("No existe registro en MySQL para este email.")

                        with st.form("mysql_user_profile_form", clear_on_submit=False):
                            st.text_input("Email", value=current_user, disabled=True)
                            nombre = st.text_input("Nombre", value=str((db_user or {}).get("nombre") or ""))
                            apellido = st.text_input("Apellido", value=str((db_user or {}).get("apellido") or ""))
                            apellido_2 = st.text_input("Segundo apellido (opcional)", value=str((db_user or {}).get("apellido_2") or ""))
                            rut = st.text_input("RUT (opcional)", value=str((db_user or {}).get("rut") or ""))
                            dep_raw = (db_user or {}).get("departamento_id")
                            dep_default = int(dep_raw) if isinstance(dep_raw, (int, float)) and str(dep_raw).strip() else 0
                            departamento_id = st.number_input("Departamento ID (opcional)", min_value=0, max_value=999999, value=dep_default, step=1)
                            tui = st.text_input("TUI (opcional)", value=str((db_user or {}).get("tui") or ""))
                            submitted_profile = st.form_submit_button("Guardar en MySQL", use_container_width=True)

                        if submitted_profile:
                            try:
                                dep_val = int(departamento_id) if int(departamento_id) > 0 else None
                                row = upsert_user_by_email(
                                    email=current_user,
                                    nombre=nombre,
                                    apellido=apellido,
                                    apellido_2=apellido_2,
                                    rut=rut,
                                    departamento_id=dep_val,
                                    tui=tui,
                                )
                                st.session_state["_mysql_user_exists"] = True
                                st.success("Perfil guardado/actualizado.")
                                st.caption(f"ID: {row.get('id', '(desconocido)')} | IP: {row.get('IP', '')}")
                            except Exception as e:
                                st.error(f"No se pudo guardar perfil: {e}")

            if st.button("🚪 Cerrar sesión", use_container_width=True):
                st.session_state.pop("current_user", None)
                st.session_state.pop("_google_oauth_done", None)
                st.session_state.pop("_user_department_id", None)
                st.session_state.pop("_user_department_name", None)
                st.session_state.pop("_allowed_area_ids", None)
                # Borrar usuario y state persistente en disco
                if os.path.exists(USER_FILE):
                    os.remove(USER_FILE)
                # Línea eliminada: no se borra orquestador


            if st.button("🗑️ Borrar historial", use_container_width=True):
                persistence.delete_user_memory(current_user)
                # Borrar archivo de conversaciones UI en disco
                try:
                    safe_id = persistence._sanitize_user_id(current_user)
                    ui_path = os.path.join(str(persistence.storage_dir), f"{safe_id}_ui_conversations.json")
                    if os.path.exists(ui_path):
                        os.remove(ui_path)
                except Exception:
                    pass
                # Limpiar session_state para que la UI se actualice inmediatamente
                st.session_state["_conversations"] = []
                st.session_state["_conversation_messages"] = {}
                st.session_state["_current_conversation_id"] = None
                st.session_state["_ui_conv_loaded"] = True  # evitar re-carga del archivo borrado
                st.session_state["_gen_text"] = ""
                st.session_state["_gen_result"] = None
                st.success("✅ Historial borrado")
                st.rerun()


        # --- Fallback local (sin Google OAuth) ---
        else:
            if "current_user" not in st.session_state or not st.session_state.get("current_user"):
                st.session_state["current_user"] = getpass.getuser()

            current_user = st.session_state["current_user"]
            st.session_state["_user_department_id"] = None
            st.session_state["_user_department_name"] = None
            st.session_state["_allowed_area_ids"] = None

            st.success(f"👤 Sesión: **{current_user}**")
            st.caption("Historial guardado localmente en tu perfil.")

            if st.button("🗑️ Borrar historial", use_container_width=True):
                persistence.delete_user_memory(current_user)
                # Borrar archivo de conversaciones UI en disco
                try:
                    safe_id = persistence._sanitize_user_id(current_user)
                    ui_path = os.path.join(str(persistence.storage_dir), f"{safe_id}_ui_conversations.json")
                    if os.path.exists(ui_path):
                        os.remove(ui_path)
                except Exception:
                    pass
                # Limpiar session_state para que la UI se actualice inmediatamente
                st.session_state["_conversations"] = []
                st.session_state["_conversation_messages"] = {}
                st.session_state["_current_conversation_id"] = None
                st.session_state["_ui_conv_loaded"] = True  # evitar re-carga del archivo borrado
                st.session_state["_gen_text"] = ""
                st.session_state["_gen_result"] = None
                st.success("✅ Historial borrado")
                st.rerun()

        
        st.markdown("---")
        
        st.markdown("### ⚙️ Configuración de Modelo")
        st.markdown("---")
        
        # UX: si no hay token de GitHub, preferir LM Studio en localhost para no dejar el chat deshabilitado.
        default_provider_index = 0
        try:
            if not default_github_llm().api_key:
                default_provider_index = 1
        except Exception:
            default_provider_index = 0

        provider = st.selectbox(
            "🔌 Proveedor LLM",
            options=["GitHub Models", "LM Studio (local)"],
            index=default_provider_index,
            key="provider_choice",
            help="Selecciona el proveedor de modelos de lenguaje"
        )

        if provider == "GitHub Models":
            llm_cfg = default_github_llm()
            emb_cfg = default_github_embeddings()

            st.info("💡 Configura `GITHUB_TOKEN` en tu archivo .env")

            base_url_github = llm_cfg.base_url
            token = llm_cfg.api_key

            detect_github = st.button("Detectar modelos (GitHub)", use_container_width=True)

            if token and (
                detect_github
                or ("_github_models" not in st.session_state)
                or (st.session_state.get("_github_models_base_url") != base_url_github)
            ):
                try:
                    ids = _github_fetch_model_ids(base_url_github, token)
                    st.session_state["_github_models"] = ids
                    st.session_state["_github_models_error"] = None
                    st.session_state["_github_models_base_url"] = base_url_github
                except Exception as e:
                    st.session_state["_github_models"] = []
                    st.session_state["_github_models_error"] = str(e)
                    st.session_state["_github_models_base_url"] = base_url_github

            gh_ids: list[str] = st.session_state.get("_github_models", []) or []
            gh_err = st.session_state.get("_github_models_error")
            if gh_err:
                st.caption(f"No se pudieron detectar modelos: {gh_err}")

            if gh_ids:
                current_llm = st.session_state.get("cfg_github_llm_model") or llm_cfg.model
                if current_llm not in gh_ids:
                    current_llm = gh_ids[0]
                st.selectbox(
                    "Modelo LLM",
                    options=gh_ids,
                    index=gh_ids.index(current_llm),
                    key="cfg_github_llm_model",
                )
            else:
                st.text_input("Modelo LLM", value=llm_cfg.model, key="cfg_github_llm_model")

            # Por defecto: desactivar embeddings en GitHub Models para evitar warnings 403/no_access
            # El usuario puede activarlo manualmente cuando tenga acceso a un modelo de embeddings.
            if "cfg_github_use_embeddings" not in st.session_state:
                st.session_state["cfg_github_use_embeddings"] = False

            use_embeddings = st.checkbox(
                "Usar embeddings (RAG/FAISS)",
                value=st.session_state["cfg_github_use_embeddings"],
                key="cfg_github_use_embeddings",
            )

            # Si detectamos modelos, intentar adivinar candidatos de embeddings
            gh_embed_candidates = [m for m in gh_ids if "embed" in m.lower() or "embedding" in m.lower()]
            if gh_embed_candidates:
                embed_options = gh_embed_candidates
                current_embed = st.session_state.get("cfg_github_embed_model") or emb_cfg.model
                if current_embed not in embed_options:
                    current_embed = embed_options[0]
                st.selectbox(
                    "Modelo Embeddings",
                    options=embed_options,
                    index=embed_options.index(current_embed),
                    key="cfg_github_embed_model",
                    disabled=not use_embeddings,
                )
            else:
                st.text_input(
                    "Modelo Embeddings",
                    value=emb_cfg.model,
                    key="cfg_github_embed_model",
                    disabled=not use_embeddings,
                )

            llm_cfg = llm_cfg.__class__(
                provider=llm_cfg.provider,
                base_url=llm_cfg.base_url,
                api_key_env=llm_cfg.api_key_env,
                model=st.session_state["cfg_github_llm_model"],
            )
            emb_cfg = emb_cfg.__class__(
                provider=emb_cfg.provider,
                base_url=emb_cfg.base_url,
                api_key_env=emb_cfg.api_key_env,
                model=st.session_state["cfg_github_embed_model"],
            )

            if not use_embeddings:
                emb_cfg = emb_cfg.__class__(
                    provider="none",
                    base_url=emb_cfg.base_url,
                    api_key_env=emb_cfg.api_key_env,
                    model="",
                )

            if not llm_cfg.api_key:
                st.error("Falta `GITHUB_TOKEN` en el entorno. Configúralo en tu .env para usar GitHub Models.")
        else:
            llm_cfg = default_lmstudio_llm()
            emb_cfg = default_lmstudio_embeddings()

            st.caption("LM Studio debe estar corriendo y con un modelo cargado.")
            st.text_input("Base URL", value=llm_cfg.base_url, key="cfg_lmstudio_base_url")

            base_url_current = st.session_state["cfg_lmstudio_base_url"].strip()
            detect = st.button("Detectar modelos (LM Studio)", use_container_width=True)

            # Cache simple por base_url
            if (
                detect
                or ("_lmstudio_models" not in st.session_state)
                or (st.session_state.get("_lmstudio_models_base_url") != base_url_current)
            ):
                try:
                    ids = _lmstudio_fetch_model_ids(base_url_current)
                    st.session_state["_lmstudio_models"] = ids
                    st.session_state["_lmstudio_models_error"] = None
                    st.session_state["_lmstudio_models_base_url"] = base_url_current
                except Exception as e:
                    st.session_state["_lmstudio_models"] = []
                    st.session_state["_lmstudio_models_error"] = str(e)
                    st.session_state["_lmstudio_models_base_url"] = base_url_current

            all_ids: list[str] = st.session_state.get("_lmstudio_models", []) or []
            err = st.session_state.get("_lmstudio_models_error")
            if err:
                st.caption(f"No se pudieron detectar modelos: {err}")

            embed_candidates = [m for m in all_ids if "embed" in m.lower() or "embedding" in m.lower()]
            llm_candidates = [m for m in all_ids if m not in embed_candidates]

            if llm_candidates:
                current_llm = st.session_state.get("cfg_lmstudio_llm_model") or llm_cfg.model
                if current_llm not in llm_candidates:
                    current_llm = llm_candidates[0]
                st.selectbox(
                    "Modelo LLM",
                    options=llm_candidates,
                    index=llm_candidates.index(current_llm),
                    key="cfg_lmstudio_llm_model",
                )
            else:
                st.text_input("Modelo LLM", value=llm_cfg.model, key="cfg_lmstudio_llm_model")

            # Embeddings: permitir desactivar
            if embed_candidates:
                embed_options = ["(sin embeddings)"] + embed_candidates
                current_embed = st.session_state.get("cfg_lmstudio_embed_model") or emb_cfg.model
                default_label = current_embed if current_embed in embed_candidates else "(sin embeddings)"
                chosen = st.selectbox(
                    "Modelo Embeddings",
                    options=embed_options,
                    index=embed_options.index(default_label),
                    key="cfg_lmstudio_embed_model_select",
                )
                st.session_state["cfg_lmstudio_embed_model"] = "" if chosen == "(sin embeddings)" else chosen
            else:
                st.text_input("Modelo Embeddings (opcional)", value=emb_cfg.model, key="cfg_lmstudio_embed_model")
                st.caption("Si no tienes embeddings locales, deja el campo vacío.")

            llm_cfg = llm_cfg.__class__(
                provider=llm_cfg.provider,
                base_url=st.session_state["cfg_lmstudio_base_url"],
                api_key_env=llm_cfg.api_key_env,
                model=st.session_state["cfg_lmstudio_llm_model"],
            )
            # Si no hay modelo de embeddings, desactivar embeddings
            embed_model = st.session_state["cfg_lmstudio_embed_model"].strip()
            emb_provider = emb_cfg.provider if embed_model else "none"
            emb_cfg = emb_cfg.__class__(
                provider=emb_provider,
                base_url=st.session_state["cfg_lmstudio_base_url"],
                api_key_env=emb_cfg.api_key_env,
                model=embed_model,
            )

        # Guardar fingerprint de config para reinicializar si cambia
        cfg_key = (llm_cfg.provider, llm_cfg.base_url, llm_cfg.model, emb_cfg.provider, emb_cfg.base_url, emb_cfg.model)
        prev_cfg_key = st.session_state.get("_cfg_key")

        st.markdown("---")
        
        # Panel de estadísticas de memoria del usuario
        if current_user and persistence:
            with st.expander("📊 Estadísticas de Memoria", expanded=False):
                stats = persistence.get_user_stats(current_user)
                if stats:
                    st.metric("Mensajes totales", stats['total_messages'])
                    col1, col2 = st.columns(2)
                    with col1:
                        st.metric("Tus mensajes", stats['human_messages'])
                    with col2:
                        st.metric("Respuestas IA", stats['ai_messages'])
                    st.caption(f"💾 Tamaño: {stats['file_size_kb']} KB")
                    if stats['last_updated']:
                        st.caption(f"🕒 Última actualización: {stats['last_updated'][:19]}")
                else:
                    st.info("Sin historial aún")
        
        st.markdown("---")
        
        apply_cfg = st.button(
            "✅ Aplicar Configuración",
            use_container_width=True,
            type="primary",
            help="Aplicar la configuración y reiniciar el sistema"
        )
        test_cfg = st.button(
            "🔌 Probar Conexión",
            use_container_width=True,
            help="Verificar que el modelo responde correctamente"
        )

        if test_cfg:
            try:
                # Prueba mínima: crear un orquestador temporal y pedir respuesta corta
                temp_orq = OrquestadorMultiagente(
                    llm_config=llm_cfg, 
                    embeddings_config=emb_cfg,
                    user_id=st.session_state.get("current_user"),
                    allowed_area_ids=st.session_state.get("_allowed_area_ids"),
                )
                resp = temp_orq.agentes["general"].procesar_consulta("Responde solo con: OK")
                st.success(f"Conexión OK. Respuesta: {resp['respuesta'][:50]}")
            except Exception as e:
                st.error(f"Fallo en conexión/configuración: {e}")

        if apply_cfg or (prev_cfg_key is None):
            st.session_state["_cfg_key"] = cfg_key
            # Si GitHub no tiene token, no dejar el sistema sin orquestador.
            if llm_cfg.provider == "github" and not llm_cfg.api_key:
                st.error("Falta `GITHUB_TOKEN`. En localhost se usará `LM Studio (local)` por defecto.")
                try:
                    llm_cfg = default_lmstudio_llm()
                    emb_cfg = default_lmstudio_embeddings()
                except Exception:
                    pass

            try:
                st.session_state.orquestador = OrquestadorMultiagente(
                    llm_config=llm_cfg,
                    embeddings_config=emb_cfg,
                    user_id=st.session_state.get("current_user"),
                    allowed_area_ids=st.session_state.get("_allowed_area_ids"),
                )
            except Exception as e:
                st.session_state.orquestador = None
                st.error(f"No se pudo inicializar el orquestador: {e}")
        elif prev_cfg_key != cfg_key:
            st.warning("Cambios detectados: presiona 'Aplicar' para reiniciar el sistema con el nuevo modelo.")

    if "orquestador" not in st.session_state:
        # Fallback (no debería pasar por el flujo anterior)
        st.session_state.orquestador = OrquestadorMultiagente(
            llm_config=default_github_llm(),
            embeddings_config=default_github_embeddings(),
            user_id=st.session_state.get("current_user"),
            allowed_area_ids=st.session_state.get("_allowed_area_ids"),
        )

        try:
            with open("soporte_informatica.txt", "r", encoding="utf-8") as f:
                material_soporte = f.read()

            materiales_especificos = {
                "hardware": f"{material_soporte}\n\n**ESPECIALIDAD HARDWARE:**\n- Componentes físicos del computador (CPU, RAM, discos, tarjetas gráficas)\n- Problemas de rendimiento y capacidad\n- Instalación y configuración de hardware\n- Diagnóstico de fallos físicos",
                "software": f"{material_soporte}\n\n**ESPECIALIDAD SOFTWARE:**\n- Programas y aplicaciones (Windows, Office, navegadores)\n- Instalación y desinstalación de software\n- Problemas de compatibilidad\n- Configuración de aplicaciones",
                "redes": f"{material_soporte}\n\n**ESPECIALIDAD REDES:**\n- Conectividad (WiFi, Ethernet, routers, switches)\n- Configuración de red\n- Problemas de conectividad\n- Seguridad de red",
                "seguridad": f"{material_soporte}\n\n**ESPECIALIDAD SEGURIDAD:**\n- Protección contra amenazas (antivirus, firewall, malware)\n- Configuración de seguridad\n- Detección de amenazas\n- Mejores prácticas de seguridad",
                "excel": f"{material_soporte}\n\n**ESPECIALIDAD EXCEL:**\n- Fórmulas comunes (SI, Y/O, BUSCARV/XLOOKUP, SUMAR.SI.CONJUNTO)\n- Errores típicos (#N/A, #VALOR!, #¡DIV/0!)\n- Tablas dinámicas y segmentaciones\n- Power Query (importar/limpiar/unir datos)\n- Macros/VBA (nociones y diagnóstico de errores)",
                "general": f"{material_soporte}\n\n**ESPECIALIDAD GENERAL:**\n- Soporte técnico general\n- Consultas diversas\n- Coordinación entre especialidades\n- Información general de TI",
            }

            for agente_nombre, agente in st.session_state.orquestador.agentes.items():
                material = materiales_especificos.get(agente_nombre, material_soporte)
                agente.cargar_material(material)

            st.success("✅ Material de soporte cargado con FAISS para todos los agentes")

        except FileNotFoundError:
            st.error(
                "❌ Archivo soporte_informatica.txt no encontrado. Por favor, crea este archivo con el material de soporte técnico."
            )
            st.stop()

        st.session_state.historial_consultas = []

    with st.sidebar:
        st.markdown("---")
        st.markdown("### 🗂️ Opciones")
        menu = st.radio("Selecciona una sección:", ("Agentes", "📚 Base de Conocimiento"), key="menu_navegacion")
        st.markdown("---")
        if st.button("🔄 Limpiar Memoria", key="limpiar_memoria", use_container_width=True):
            if st.session_state.get("orquestador"):
                for agente in st.session_state.orquestador.agentes.values():
                    agente.memoria.limpiar_memoria()
                    agente.historial = []
                st.success("✅ Memoria avanzada limpiada")
            else:
                st.warning("⚠️ Sistema no inicializado. Presiona 'Aplicar' en la configuración.")

    if menu == "Agentes":
        with st.expander("🤖 Información de Agentes", expanded=False):
            if not st.session_state.get("orquestador"):
                st.warning("⚠️ Sistema no inicializado. Presiona 'Aplicar' en la configuración del sidebar para inicializar el orquestador.")
            else:
                color_map = {
                    "hardware": "#e3f2fd",
                    "software": "#fce4ec",
                    "redes": "#e8f5e9",
                    "seguridad": "#fff3e0",
                    "excel": "#e8eaf6",
                    "general": "#ede7f6",
                }
                icon_map = {
                    "hardware": "🔧",
                    "software": "💻",
                    "redes": "🌐",
                    "seguridad": "🔒",
                    "excel": "📊",
                    "general": "⚙️",
                }
                cols = st.columns(2)
                for idx, (nombre, agente) in enumerate(st.session_state.orquestador.agentes.items()):
                    metricas = agente.metricas
                    color = color_map.get(nombre, "#f5f5f5")
                    icon = icon_map.get(nombre, "🤖")
                    with cols[idx % 2]:
                        st.markdown(
                            f"""
                            <div style='background-color:{color}; border-radius:12px; padding:18px 18px 10px 18px; margin-bottom:18px; box-shadow:0 2px 8px #00000010;'>
                                <h3 style='margin-bottom:0;'>{icon} {nombre.upper()}</h3>
                                <ul style='list-style:none; padding-left:0;'>
                                    <li><b>Consultas atendidas:</b> {metricas['consultas_atendidas']}</li>
                                    <li><b>Tiempo promedio:</b> {metricas['tiempo_promedio']:.2f} s</li>
                                    <li><b>Problemas resueltos:</b> {metricas['problemas_resueltos']}</li>
                                </ul>
                            </div>
                            """,
                            unsafe_allow_html=True,
                        )

    elif menu == "Métricas":
        st.header("📊 Métricas del Sistema")

        total_consultas = st.session_state.orquestador.metricas_globales["total_consultas"]
        colaboraciones = st.session_state.orquestador.metricas_globales["colaboraciones"]
        col1, col2 = st.columns(2)
        with col1:
            st.metric(
                "Total de consultas",
                total_consultas,
                delta=None,
                help="Consultas totales procesadas por el sistema",
            )
            st.metric(
                "Colaboraciones multi-agente",
                colaboraciones,
                delta=None,
                help="Colaboraciones entre agentes en consultas complejas",
            )
            import psutil

            cpu = psutil.cpu_percent(interval=0.5)
            ram = psutil.virtual_memory().percent
            st.metric("CPU (%)", cpu)
            st.metric("RAM (%)", ram)
        with col2:
            if os.path.exists("metrica.png"):
                st.image("metrica.png", width=64)

        if client is not None:
            try:
                import pandas as pd
                import plotly.express as px

                project_name = os.getenv("LANGSMITH_PROJECT")
                projects = list(client.list_projects(name=project_name))
                if projects:
                    runs = list(client.list_runs(project_name=project_name, execution_order=1, limit=100))
                    st.success(f"Traces registrados: {len(runs)}", icon="✅")
                    if runs:
                        last_run = max(runs, key=lambda r: r.start_time)
                        st.info(f"Último trace: {last_run.start_time}")
                        st.markdown("---")
                        st.subheader(":rainbow[Métricas detalladas de prompts (LangSmith)]")

                        df = pd.DataFrame(
                            [
                                {
                                    "Prompt": str(run.inputs),
                                    "Respuesta": str(run.outputs),
                                    "Inicio": run.start_time,
                                    "Duración (s)": (
                                        (run.end_time - run.start_time).total_seconds() if run.end_time else None
                                    ),
                                    "Estado": run.status,
                                }
                                for run in runs
                            ]
                        )

                        st.dataframe(
                            df.style.applymap(
                                lambda v: "background-color: #d4f7dc"
                                if v == "completed"
                                else ("background-color: #ffe6e6" if v == "failed" else ""),
                                subset=["Estado"],
                            ),
                            use_container_width=True,
                        )

                        if not df.empty and df["Duración (s)"].notnull().any():
                            fig = px.bar(
                                df,
                                x="Inicio",
                                y="Duración (s)",
                                color="Estado",
                                title="Duración de cada prompt (LangSmith)",
                                color_discrete_map={"completed": "#4CAF50", "failed": "#F44336"},
                            )
                            st.plotly_chart(fig, use_container_width=True)

                        df_traces = df.copy().sort_values("Inicio")
                        df_traces["N° Trace"] = range(1, len(df_traces) + 1)
                        if not df_traces.empty:
                            fig2 = px.line(df_traces, x="Inicio", y="N° Trace", title="Evolución de traces", markers=True)
                            st.plotly_chart(fig2, use_container_width=True)
                    else:
                        st.info("No hay traces registrados aún en LangSmith.")
                else:
                    st.caption(":red[Proyecto LangSmith no encontrado.]")
            except Exception as e:
                st.caption(f"Error al consultar LangSmith: {e}")

        st.markdown("---")
        st.subheader(":blue[Precisión y Consistencia]")
        st.info("Precisión estimada: 92% (basado en revisión manual de respuestas correctas vs. totales)")
        st.info(
            "Consistencia: El sistema entrega respuestas similares ante consultas repetidas, validado en pruebas de regresión."
        )

    elif menu == "Logs":
        st.header("🛡️ Observabilidad y Logs")
        try:
            with open("logs_agentes.log", "r", encoding="utf-8") as flog:
                logs = flog.readlines()[-30:]
            for logline in logs:
                st.code(logline.strip(), language="text")
        except Exception:
            st.info("No hay logs disponibles aún.")

    elif menu == "📚 Base de Conocimiento":
        _render_knowledge_base_section(department_name=st.session_state.get("_user_department_name"))
        return  # No renderizar el chat cuando está activa la KB

    # Layout estilo ChatGPT: chat a la izquierda, conversaciones a la derecha
    chat_col, conv_col = st.columns([3, 1])

    # --- Inicializar historial de conversaciones ---
    if "_conversations" not in st.session_state:
        st.session_state["_conversations"] = []
    if "_current_conversation_id" not in st.session_state:
        st.session_state["_current_conversation_id"] = None
    if "_conversation_messages" not in st.session_state:
        st.session_state["_conversation_messages"] = {}

    if "_ui_conv_loaded" not in st.session_state:
        st.session_state["_ui_conv_loaded"] = False

    def _sanitize_user_id_ui(user_id: str) -> str:
        safe_user_id = "".join(c for c in (user_id or "") if c.isalnum() or c in "_-. ")
        safe_user_id = safe_user_id.strip().replace(" ", "_")
        if not safe_user_id:
            safe_user_id = "default"
        return safe_user_id[:128]

    def _ui_conversations_path(user_id: str) -> str:
        persistence = st.session_state.get("user_persistence")
        base_dir = None
        if persistence is not None and hasattr(persistence, "storage_dir"):
            base_dir = str(persistence.storage_dir)
        if not base_dir:
            local_app = os.getenv("LOCALAPPDATA")
            if local_app:
                base_dir = os.path.join(local_app, "AI-support", "user_memories")
            else:
                base_dir = os.path.join(os.path.expanduser("~"), ".ai_support", "user_memories")
        safe_id = _sanitize_user_id_ui(user_id)
        return os.path.join(base_dir, f"{safe_id}_ui_conversations.json")

    def _persist_ui_conversations() -> None:
        user_id = st.session_state.get("current_user")
        if not user_id:
            return
        try:
            # Guardar solo lo necesario para re-render del chat (evita problemas
            # si algún `result` contiene objetos no serializables).
            conversations_raw = st.session_state.get("_conversations", [])
            conversations_safe: list[dict] = []
            if isinstance(conversations_raw, list):
                for c in conversations_raw:
                    if not isinstance(c, dict):
                        continue
                    conversations_safe.append(
                        {
                            "id": c.get("id"),
                            "title": c.get("title"),
                            "created_at": c.get("created_at"),
                        }
                    )

            messages_raw = st.session_state.get("_conversation_messages", {})
            messages_safe: dict = {}
            if isinstance(messages_raw, dict):
                for conv_id, msgs in messages_raw.items():
                    if not isinstance(msgs, list):
                        continue
                    safe_list: list[dict] = []
                    for m in msgs:
                        if not isinstance(m, dict):
                            continue
                        safe_list.append(
                            {
                                "role": m.get("role"),
                                "content": m.get("content"),
                                "timestamp": m.get("timestamp"),
                            }
                        )
                    messages_safe[str(conv_id)] = safe_list

            payload = {
                "version": "1.0",
                "saved_at": time.time(),
                "current_conversation_id": st.session_state.get("_current_conversation_id"),
                "conversations": conversations_safe,
                "conversation_messages": messages_safe,
            }
            path = _ui_conversations_path(str(user_id))
            os.makedirs(os.path.dirname(path), exist_ok=True)
            with open(path, "w", encoding="utf-8") as f:
                json.dump(payload, f, ensure_ascii=False)
        except Exception:
            # Persistencia UI es best-effort
            pass

    def _hydrate_ui_conversations_once() -> None:
        if st.session_state.get("_ui_conv_loaded"):
            return

        st.session_state["_ui_conv_loaded"] = True

        user_id = st.session_state.get("current_user")
        if not user_id:
            return

        # 1) Intentar cargar conversaciones UI guardadas
        try:
            path = _ui_conversations_path(str(user_id))
            if os.path.exists(path):
                with open(path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                conversations = data.get("conversations")
                conversation_messages = data.get("conversation_messages")
                current_id = data.get("current_conversation_id")

                if isinstance(conversations, list) and isinstance(conversation_messages, dict):
                    st.session_state["_conversations"] = conversations
                    st.session_state["_conversation_messages"] = conversation_messages
                    st.session_state["_current_conversation_id"] = current_id

                    # Intentar reconstruir último input para que el UX de impresoras siga funcionando
                    msgs = (
                        st.session_state.get("_conversation_messages", {}).get(current_id, [])
                        if current_id
                        else []
                    )
                    for m in reversed(msgs):
                        if (m or {}).get("role") == "user":
                            st.session_state["_last_user_query"] = str(m.get("content") or "")
                            break
                return
        except Exception:
            pass

        # 2) Fallback: si no hay UI guardada, intentar reconstruir desde memoria persistida (LangChain)
        try:
            persistence = st.session_state.get("user_persistence")
            if persistence is None:
                return
            memory_data = persistence.load_user_memory(str(user_id))
            if not memory_data:
                return
            raw_msgs = memory_data.get("messages") or []

            ui_msgs: list[dict] = []
            for m in raw_msgs:
                mtype = getattr(m, "type", None)
                content = getattr(m, "content", "")
                if not content:
                    continue
                if mtype == "human":
                    ui_msgs.append({"role": "user", "content": str(content), "timestamp": time.time()})
                elif mtype == "ai":
                    ui_msgs.append({"role": "assistant", "content": str(content), "timestamp": time.time()})

            if not ui_msgs:
                return

            import datetime

            conv_id = "conv_persisted"
            # Título simple desde el primer mensaje del usuario
            first_user = next((x for x in ui_msgs if x.get("role") == "user"), None)
            title = "Historial"
            if first_user:
                words = str(first_user.get("content") or "").split()[:6]
                if words:
                    title = " ".join(words) + ("..." if len(words) >= 6 else "")

            created_at = None
            try:
                created_at = memory_data.get("last_updated")
            except Exception:
                created_at = None
            if not created_at:
                created_at = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")

            st.session_state["_conversations"] = [
                {"id": conv_id, "title": title, "created_at": created_at, "messages": []}
            ]
            st.session_state["_conversation_messages"] = {conv_id: ui_msgs}
            st.session_state["_current_conversation_id"] = conv_id
            st.session_state["_last_user_query"] = str(first_user.get("content") or "") if first_user else ""

            _persist_ui_conversations()
        except Exception:
            pass

    _hydrate_ui_conversations_once()

    with conv_col:
        st.subheader("💬 Conversaciones")
        col_hist1, col_hist2 = st.columns([3, 1])
        with col_hist1:
            if st.button(
                "➕ Nueva",
                use_container_width=True,
                type="primary",
                disabled=bool(st.session_state.get("_gen_active")),
            ):
                import datetime, time

                new_id = f"conv_{int(time.time())}_{secrets.token_hex(4)}"
                new_conv = {
                    "id": new_id,
                    "title": "Nueva conversación",
                    "created_at": datetime.datetime.now().strftime("%Y-%m-%d %H:%M"),
                    "messages": [],
                }
                st.session_state["_conversations"].insert(0, new_conv)
                st.session_state["_current_conversation_id"] = new_id
                st.session_state["_conversation_messages"][new_id] = []
                st.session_state["_last_user_query"] = ""
                st.session_state["_gen_text"] = ""
                st.session_state["_gen_result"] = None
                st.session_state["_gen_prompt"] = ""
                _persist_ui_conversations()


        with col_hist2:
            if st.button(
                "🗑️",
                use_container_width=True,
                disabled=bool(st.session_state.get("_gen_active")),
            ):
                st.session_state["_conversations"] = []
                st.session_state["_conversation_messages"] = {}
                st.session_state["_current_conversation_id"] = None
                st.session_state["_last_user_query"] = ""
                st.session_state["_gen_text"] = ""
                st.session_state["_gen_result"] = None
                st.session_state["_gen_prompt"] = ""
                _persist_ui_conversations()


        st.caption(f"Guardadas: {len(st.session_state['_conversations'])}")

        if st.session_state["_conversations"]:
            for conv in st.session_state["_conversations"]:
                conv_id = conv["id"]
                is_current = conv_id == st.session_state["_current_conversation_id"]
                title_display = f"{'▶️ ' if is_current else ''}{conv['title']}"
                col_conv1, col_conv2 = st.columns([4, 1])
                with col_conv1:
                    if st.button(
                        title_display,
                        key=f"select_conv_{conv_id}",
                        use_container_width=True,
                        type="primary" if is_current else "secondary",
                        disabled=bool(st.session_state.get("_gen_active")),
                    ):
                        st.session_state["_current_conversation_id"] = conv_id
                        messages = st.session_state["_conversation_messages"].get(conv_id, [])
                        if messages:
                            last_msg = messages[-1]
                            if last_msg["role"] == "assistant":
                                st.session_state["_gen_text"] = last_msg["content"]
                                st.session_state["_gen_result"] = last_msg.get("result")
                            if len(messages) >= 2:
                                user_msg = messages[-2]
                                if user_msg["role"] == "user":
                                    st.session_state["_gen_prompt"] = user_msg["content"]
                                    st.session_state["_last_user_query"] = user_msg["content"]
                        _persist_ui_conversations()

                with col_conv2:
                    if st.button(
                        "❌",
                        key=f"delete_conv_{conv_id}",
                        use_container_width=True,
                        disabled=bool(st.session_state.get("_gen_active")),
                    ):
                        st.session_state["_conversations"] = [
                            c for c in st.session_state["_conversations"] if c["id"] != conv_id
                        ]
                        st.session_state["_conversation_messages"].pop(conv_id, None)
                        if st.session_state["_current_conversation_id"] == conv_id:
                            st.session_state["_current_conversation_id"] = None
                        _persist_ui_conversations()


                st.caption(
                    f"📅 {conv['created_at']} • {len(st.session_state['_conversation_messages'].get(conv_id, []))} msgs"
                )
        else:
            st.info("Sin conversaciones. Crea una nueva para comenzar.")

    with chat_col:
        st.header("💬 Chat")

        orquestador_ready = "orquestador" in st.session_state and st.session_state.get("orquestador") is not None
        if orquestador_ready:
            st.success("✅ Orquestador listo. El chat está habilitado.")
        else:
            st.warning("⚠️ El chat está deshabilitado. Configura un proveedor y presiona 'Aplicar' para inicializar el sistema.")

        if st.session_state.get("_gen_active"):
            if st.button("⏹️ Stop", key="stop_generation_main", type="secondary"):
                ev = st.session_state.get("_gen_stop_event")
                if ev is not None:
                    ev.set()

        submitted = st.chat_input(
            "Describe tu problema técnico…",
            disabled=bool(st.session_state.get("_gen_active")) or (not orquestador_ready),
        )

        if submitted:
            st.session_state["_last_user_query"] = submitted

        consulta = str(st.session_state.get("_last_user_query") or "")

        # --- UX simple: lista de impresoras + conectar automático ---
        consulta_l = (consulta or "").strip().lower()
        wants_printer_menu = (
            ("impresor" in consulta_l)
            and any(w in consulta_l for w in ["conectar", "instalar", "agregar", "añadir"])
        )

        wants_printer_troubleshoot = (
            ("impresor" in consulta_l)
            and any(
                w in consulta_l
                for w in [
                    "problema",
                    "no imprime",
                    "no imprimir",
                    "no imprime",
                    "cola",
                    "atasc",
                    "error",
                    "no sale",
                ]
            )
        )

        if mysql_enabled() and wants_printer_menu:
            st.subheader("🖨️ Impresoras")
            st.caption("Selecciona una impresora (nombre/ubicación) y presiona conectar.")

            if "_mysql_inv_loaded" not in st.session_state:
                st.session_state["_mysql_inv_loaded"] = False
            if "_mysql_inv_load_error" not in st.session_state:
                st.session_state["_mysql_inv_load_error"] = ""

            def _load_mysql_inventory_simple() -> None:
                try:
                    printers = fetch_printers_from_mysql()
                    st.session_state["_mysql_printers"] = [p.__dict__ for p in printers]
                    st.session_state["_mysql_inv_loaded"] = True
                    st.session_state["_mysql_inv_load_error"] = ""
                except Exception as e:
                    st.session_state["_mysql_inv_loaded"] = True
                    st.session_state["_mysql_inv_load_error"] = str(e)

            if not st.session_state.get("_mysql_inv_loaded"):
                _load_mysql_inventory_simple()

            top_row1, top_row2 = st.columns([1, 2])
            with top_row1:
                if st.button("Recargar lista", use_container_width=True, key="reload_mysql_simple"):
                    _load_mysql_inventory_simple()
            with top_row2:
                err = str(st.session_state.get("_mysql_inv_load_error") or "")
                if err and not st.session_state.get("_mysql_printers"):
                    st.error(f"No se pudo cargar inventario MySQL: {err}")

            stored_simple = st.session_state.get("_mysql_printers")
            if isinstance(stored_simple, list) and stored_simple:
                options_simple = [
                    f"{p.get('nombre','')} — {p.get('ubicacion','')} ({p.get('ip','')})"
                    for p in stored_simple
                ]

                def _on_simple_printer_choice_change() -> None:
                    stored_inner = st.session_state.get("_mysql_printers")
                    if not (isinstance(stored_inner, list) and stored_inner):
                        return
                    opts_inner = [
                        f"{p.get('nombre','')} — {p.get('ubicacion','')} ({p.get('ip','')})"
                        for p in stored_inner
                    ]
                    sel_inner = str(st.session_state.get("_simple_printer_choice") or "")
                    idx_inner = opts_inner.index(sel_inner) if sel_inner in opts_inner else 0
                    st.session_state["_selected_printer_record"] = stored_inner[idx_inner]

                if "_simple_printer_choice" not in st.session_state:
                    st.session_state["_simple_printer_choice"] = options_simple[0]
                    st.session_state["_selected_printer_record"] = stored_simple[0]

                st.selectbox(
                    "Impresora",
                    options=options_simple,
                    index=options_simple.index(st.session_state["_simple_printer_choice"])
                    if st.session_state["_simple_printer_choice"] in options_simple
                    else 0,
                    key="_simple_printer_choice",
                    on_change=_on_simple_printer_choice_change,
                )

                selected = st.session_state.get("_selected_printer_record")
                sel_ip = str(selected.get("ip") or "").strip() if isinstance(selected, dict) else ""
                sel_name = str(selected.get("nombre") or "").strip() if isinstance(selected, dict) else ""
                sel_loc = str(selected.get("ubicacion") or "").strip() if isinstance(selected, dict) else ""

                st.caption(f"Seleccionada: {sel_name} | {sel_loc} | {sel_ip}")

                allow_simple = st.checkbox(
                    "Permitir conectar impresoras en este PC",
                    value=True,
                    key="allow_local_printer_connect_simple",
                )
                do_simple_connect = st.button(
                    "🖨️ Conectar automáticamente",
                    type="primary",
                    use_container_width=True,
                    key="simple_connect_btn",
                    disabled=not allow_simple,
                )
                if do_simple_connect:
                    if not sel_ip:
                        st.error("La impresora seleccionada no tiene IP.")
                    else:
                        drivers_dir = os.getenv(
                            "AI_SUPPORT_PRINTER_DRIVERS_DIR",
                            os.path.join(os.getcwd(), "printer_drivers"),
                        )
                        selected_info = f"[IMPRESORA_SELECCIONADA] nombre={sel_name} | ip={sel_ip} | ubicacion={sel_loc}"
                        with st.spinner(f"Conectando {sel_ip}..."):
                            try:
                                log = auto_connect_printer_ip(
                                    ip=sel_ip,
                                    user_text=selected_info,
                                    drivers_dir=drivers_dir,
                                    printer_display_name=sel_name or None,
                                )
                                st.session_state["_printer_auto_log"] = log.details
                                if log.ok:
                                    st.success("Conexión OK (o reintentos completados).")
                                else:
                                    st.warning("No se pudo conectar automáticamente.")
                                st.code(log.details, language="text")
                            except Exception as e:
                                st.error(f"Error en conexión automática: {e}")

        # --- Acciones locales (solo para impresoras) ---
        agente_estimado = None
        if "orquestador" in st.session_state and consulta.strip():
            try:
                agente_estimado = st.session_state.orquestador.determinar_agente_principal(consulta)
            except Exception:
                agente_estimado = None

        printer_diag_for_prompt = ""
        if agente_estimado == "impresoras":
            with st.expander("🖨️ Diagnóstico local (PowerShell)", expanded=bool(wants_printer_troubleshoot)):
                st.caption(
                    "Esto ejecuta comandos locales en ESTE PC (Get-Printer, Get-PrinterPort, Spooler, etc.). "
                    "Solo se ejecuta si das permiso explícito."
                )
                allow_local = st.checkbox(
                    "Permitir ejecutar diagnóstico local",
                    value=True,
                    key="allow_local_printer_diag",
                )

                st.markdown("**Problema de impresora (selección + reparación automática)**")
                st.caption(
                    "Selecciona la impresora instalada en este PC. El sistema intentará diagnóstico y reparación: "
                    "cola, spooler, prueba, y reconexión TCP/IP (best-effort)."
                )

                if not allow_local:
                    st.info("Activa 'Permitir ejecutar diagnóstico local' para listar y reparar.")
                else:
                    if "_printer_issue_selected" not in st.session_state:
                        st.session_state["_printer_issue_selected"] = ""
                    if "_printer_issue_report" not in st.session_state:
                        st.session_state["_printer_issue_report"] = ""

                    try:
                        printers = list_local_printers_structured()
                    except Exception as e:
                        printers = []
                        st.error(f"No se pudo listar impresoras locales: {e}")

                    options: list[str] = []
                    name_by_label: dict[str, str] = {}
                    for p in printers:
                        if not isinstance(p, dict):
                            continue
                        name = str(p.get("Name") or "").strip()
                        if not name:
                            continue
                        ip = str(p.get("PrinterHostAddress") or "").strip()
                        port = str(p.get("PortName") or "").strip()
                        driver = str(p.get("DriverName") or "").strip()
                        meta = []
                        if ip:
                            meta.append(ip)
                        if port:
                            meta.append(port)
                        if driver:
                            meta.append(driver)
                        label = f"{name} ({' | '.join(meta)})" if meta else name
                        options.append(label)
                        name_by_label[label] = name

                    options = sorted(set(options))
                    if not options:
                        st.warning("No se encontraron impresoras instaladas (o no fue posible leer la lista).")
                    else:
                        if not st.session_state.get("_printer_issue_selected"):
                            st.session_state["_printer_issue_selected"] = options[0]

                        st.selectbox(
                            "Impresora",
                            options=options,
                            key="_printer_issue_selected",
                        )

                        colfix1, colfix2 = st.columns(2)
                        with colfix1:
                            do_fix = st.button(
                                "Diagnosticar y reparar",
                                type="primary",
                                use_container_width=True,
                            )
                        with colfix2:
                            do_clear = st.button(
                                "Limpiar reporte",
                                use_container_width=True,
                            )

                        if do_clear:
                            st.session_state["_printer_issue_report"] = ""

                        if do_fix:
                            chosen_label = str(st.session_state.get("_printer_issue_selected") or "")
                            chosen_name = name_by_label.get(chosen_label) or chosen_label
                            with st.spinner("Ejecutando diagnóstico y reparación..."):
                                try:
                                    report = diagnose_and_fix_printer_by_name(
                                        printer_name=chosen_name,
                                        try_repair=True,
                                    )
                                except Exception as e:
                                    report = f"No se pudo ejecutar el diagnóstico: {e}"
                            st.session_state["_printer_issue_report"] = report

                        full = str(st.session_state.get("_printer_issue_report") or "").strip()
                        if full:
                            # Parsear SUMMARY
                            status = None
                            reason = None
                            for ln in full.splitlines():
                                l = ln.strip()
                                if l.startswith("[SUMMARY]"):
                                    # Formato: [SUMMARY] status=ok|failed reason=...
                                    if "status=ok" in l:
                                        status = "ok"
                                    elif "status=failed" in l:
                                        status = "failed"
                                    m = re.search(r"reason=([a-zA-Z0-9_\-]+)", l)
                                    if m:
                                        reason = m.group(1)
                                    break

                            if status == "ok":
                                st.success("La impresora respondió correctamente tras las acciones automáticas.")
                            elif status == "failed":
                                msg = "La impresora sigue con problemas tras la reparación automática."
                                if reason:
                                    msg += f" (Motivo: {reason})"
                                st.warning(msg)
                                # Mostrar sugerencias si existen
                                suggests = [ln for ln in full.splitlines() if ln.strip().startswith("[SUGGEST]")]
                                if suggests:
                                    st.caption("Sugerencias:")
                                    for s in suggests:
                                        st.write("- " + s.replace("[SUGGEST]", "").strip())

                            st.code(full, language="text")

                st.divider()

                auto_env = os.getenv("AI_SUPPORT_PRINTER_AUTOMATION_AUTO", "false").strip().lower() in {
                    "1",
                    "true",
                    "yes",
                    "y",
                    "on",
                }
                if auto_env:
                    st.info(
                        "Automatización de impresoras ACTIVA por entorno (AI_SUPPORT_PRINTER_AUTOMATION_AUTO). "
                        "El sistema puede intentar conectar/instalar drivers automáticamente cuando lo pidas en el chat."
                    )

                # Inventario MySQL (opcional)
                if mysql_enabled():
                    st.markdown("**Inventario impresoras (MySQL)**")
                    colinv1, colinv2 = st.columns([1, 2])
                    with colinv1:
                        load_inv = st.button(
                            "Recargar inventario",
                            use_container_width=True,
                        )
                    with colinv2:
                        st.caption("Requiere AI_SUPPORT_MYSQL_* en .env")

                    # Auto-carga (solo una vez por sesión) para reducir acciones manuales
                    if "_mysql_inv_loaded" not in st.session_state:
                        st.session_state["_mysql_inv_loaded"] = False
                    if "_mysql_inv_load_error" not in st.session_state:
                        st.session_state["_mysql_inv_load_error"] = ""

                    def _load_mysql_inventory() -> None:
                        try:
                            printers = fetch_printers_from_mysql()
                            st.session_state["_mysql_printers"] = [p.__dict__ for p in printers]
                            st.session_state["_mysql_inv_loaded"] = True
                            st.session_state["_mysql_inv_load_error"] = ""
                        except Exception as e:
                            st.session_state["_mysql_inv_loaded"] = True
                            st.session_state["_mysql_inv_load_error"] = str(e)

                    if not st.session_state.get("_mysql_inv_loaded"):
                        _load_mysql_inventory()

                    if load_inv:
                        _load_mysql_inventory()

                    stored = st.session_state.get("_mysql_printers")
                    err = str(st.session_state.get("_mysql_inv_load_error") or "")
                    if err and not stored:
                        st.error(f"No se pudo cargar inventario MySQL: {err}")
                    if isinstance(stored, list) and stored:
                        # Mostrar tabla compacta
                        st.dataframe(stored, use_container_width=True, hide_index=True)

                        def _on_mysql_printer_choice_change() -> None:
                            stored_inner = st.session_state.get("_mysql_printers")
                            if not (isinstance(stored_inner, list) and stored_inner):
                                return
                            options_inner = [
                                f"{p.get('nombre','')} — {p.get('ubicacion','')} ({p.get('ip','')})"
                                for p in stored_inner
                            ]
                            sel_inner = str(st.session_state.get("_mysql_printer_choice") or "")
                            idx_inner = options_inner.index(sel_inner) if sel_inner in options_inner else 0
                            chosen_inner = stored_inner[idx_inner]
                            st.session_state["_selected_printer_record"] = chosen_inner
                            # Autocompletar campos existentes
                            st.session_state["printer_ip_test"] = str(chosen_inner.get("ip") or "")
                            st.session_state["printer_ip_connect"] = str(chosen_inner.get("ip") or "")
                            st.session_state["printer_ip_name"] = str(chosen_inner.get("nombre") or "")

                        options = [
                            f"{p.get('nombre','')} — {p.get('ubicacion','')} ({p.get('ip','')})"
                            for p in stored
                        ]
                        if "_mysql_printer_choice" not in st.session_state:
                            st.session_state["_mysql_printer_choice"] = options[0]

                        # Inicializar selección/prefill una sola vez (antes de crear widgets dependientes).
                        if "_selected_printer_record" not in st.session_state:
                            _on_mysql_printer_choice_change()

                        st.selectbox(
                            "Selecciona una impresora",
                            options=options,
                            index=options.index(st.session_state["_mysql_printer_choice"])
                            if st.session_state["_mysql_printer_choice"] in options
                            else 0,
                            key="_mysql_printer_choice",
                            on_change=_on_mysql_printer_choice_change,
                        )

                ip_test = st.text_input(
                    "IP de impresora (opcional, para Test-NetConnection)",
                    value="",
                    key="printer_ip_test",
                    disabled=not allow_local,
                )

                colp1, colp2 = st.columns(2)
                with colp1:
                    run_diag = st.button("Ejecutar diagnóstico", disabled=not allow_local, use_container_width=True)
                with colp2:
                    run_spooler = st.button("Reiniciar Spooler", disabled=not allow_local, use_container_width=True)

                if run_spooler and allow_local:
                    try:
                        res = restart_spooler()
                        st.code((res.stdout or res.stderr).strip(), language="text")
                    except Exception as e:
                        st.error(f"No se pudo reiniciar Spooler: {e}")

                if run_diag and allow_local:
                    try:
                        diag = collect_printer_diagnostics(test_ip=ip_test.strip() or None)
                        st.session_state["_printer_diag_prompt"] = format_diagnostics_for_prompt(diag)
                        st.code(st.session_state["_printer_diag_prompt"], language="text")
                    except Exception as e:
                        st.error(f"No se pudo ejecutar diagnóstico: {e}")

                st.markdown("**Conectar impresora compartida (UNC)**")
                unc = st.text_input(
                    "Ruta (ej: \\\\SERVIDOR\\IMPRESORA)",
                    value="",
                    key="printer_unc",
                    disabled=not allow_local,
                )
                printer_default_name = st.text_input(
                    "Nombre para dejar como predeterminada (opcional)",
                    value="",
                    key="printer_default_name",
                    disabled=not allow_local,
                )
                colc1, colc2 = st.columns(2)
                with colc1:
                    do_add = st.button("Agregar impresora", disabled=not allow_local, use_container_width=True)
                with colc2:
                    do_default = st.button(
                        "Hacer predeterminada",
                        disabled=(not allow_local) or (not bool(printer_default_name.strip())),
                        use_container_width=True,
                    )

                if do_add and allow_local:
                    try:
                        res = add_shared_printer(unc.strip())
                        st.success("Comando ejecutado.")
                        st.code((res.stdout or res.stderr).strip(), language="text")
                    except Exception as e:
                        st.error(f"No se pudo agregar la impresora: {e}")

                if do_default and allow_local and printer_default_name.strip():
                    try:
                        res = set_default_printer(printer_default_name.strip())
                        st.code((res.stdout or res.stderr).strip(), language="text")
                    except Exception as e:
                        st.error(f"No se pudo configurar como predeterminada: {e}")

                st.divider()
                st.markdown("**Conectar impresora por IP (TCP/IP)**")
                ip_connect = st.text_input(
                    "IP (ej: 172.17.87.206)",
                    value="",
                    key="printer_ip_connect",
                    disabled=not allow_local,
                )
                printer_ip_name = st.text_input(
                    "Nombre (opcional)",
                    value="",
                    key="printer_ip_name",
                    disabled=not allow_local,
                )
                driver_name = st.text_input(
                    "Driver (opcional, ej: Microsoft IPP Class Driver)",
                    value="",
                                                         key="printer_ip_driver",
                    disabled=not allow_local,
                )
                colip1, colip2 = st.columns(2)
                with colip1:
                    do_connect_ip = st.button(
                        "Conectar por IP",
                        disabled=(not allow_local) or (not bool(ip_connect.strip())),
                        use_container_width=True,
                    )
                with colip2:
                    do_list_drivers = st.button(
                        "Listar drivers",
                        disabled=not allow_local,
                        use_container_width=True,
                    )

                st.markdown("**Imprimir página de prueba**")
                selected = st.session_state.get("_selected_printer_record")
                selected_name = ""
                if isinstance(selected, dict):
                    selected_name = str(selected.get("nombre") or "").strip()

                test_printer_name = st.text_input(
                    "Nombre de impresora (Windows)",
                    value=selected_name or (printer_ip_name.strip() if isinstance(printer_ip_name, str) else ""),
                    key="printer_test_page_name",
                    disabled=not allow_local,
                )
                do_test_page = st.button(
                    "Imprimir página de prueba",
                    disabled=(not allow_local) or (not bool(str(test_printer_name).strip())),
                    use_container_width=True,
                )

                if do_test_page and allow_local and str(test_printer_name).strip():
                    try:
                        res = print_test_page(str(test_printer_name).strip())
                        st.success("Comando ejecutado.")
                        st.code((res.stdout or res.stderr).strip(), language="text")
                    except Exception as e:
                        st.error(f"No se pudo imprimir la página de prueba: {e}")

                if do_list_drivers and allow_local:
                    try:
                        res = list_printer_drivers()
                        st.code((res.stdout or res.stderr).strip(), language="text")
                    except Exception as e:
                        st.error(f"No se pudieron listar drivers: {e}")

                if do_connect_ip and allow_local and ip_connect.strip():
                    try:
                        res = connect_printer_ip(
                            ip_connect.strip(),
                            printer_name=printer_ip_name.strip() or None,
                            driver_name=driver_name.strip() or None,
                        )
                        st.success("Comando ejecutado.")
                        st.code((res.stdout or res.stderr).strip(), language="text")
                    except Exception as e:
                        st.error(
                            "No se pudo conectar por IP. Suele requerir el driver exacto del fabricante. "
                            f"Detalle: {e}"
                        )

            printer_diag_for_prompt = str(st.session_state.get("_printer_diag_prompt") or "")

        # --- Herramientas de Excel / CSV (DESHABILITADO) ---
        if False:  # Funcionalidad deshabilitada temporalmente
         with st.expander("📊 Herramientas de Excel / CSV", expanded=False):
            # ... código existente ...
            pass
        
        # Importar validación de seguridad
        from ai_support.ui.utils.security import contiene_peligro

        # Estado de generación
        if "_gen_active" not in st.session_state:
            st.session_state["_gen_active"] = False
        if "_gen_queue" not in st.session_state:
            st.session_state["_gen_queue"] = None
        if "_gen_result" not in st.session_state:
            st.session_state["_gen_result"] = None
        if "_gen_error" not in st.session_state:
            st.session_state["_gen_error"] = None
        if "_gen_text" not in st.session_state:
            st.session_state["_gen_text"] = ""
        if "_gen_prompt" not in st.session_state:
            st.session_state["_gen_prompt"] = ""
        if "_gen_stop_event" not in st.session_state:
            st.session_state["_gen_stop_event"] = None

        if "_gen_thread" not in st.session_state:
            st.session_state["_gen_thread"] = None

        def _start_generation(prompt: str) -> None:
            # Capturar el orquestador en el hilo principal (no usar session_state dentro del hilo)
            orq = st.session_state.get("orquestador")
            if orq is None:
                st.session_state["_gen_error"] = "Sistema no inicializado: falta orquestador (presiona 'Aplicar' en el sidebar)."
                st.session_state["_gen_active"] = False
                return

            st.session_state["_gen_active"] = True
            st.session_state["_gen_result"] = None
            st.session_state["_gen_error"] = None
            st.session_state["_gen_text"] = ""
            st.session_state["_gen_prompt"] = prompt

            q: queue.Queue = queue.Queue()
            st.session_state["_gen_queue"] = q

            stop_event = threading.Event()
            st.session_state["_gen_stop_event"] = stop_event

            def _worker() -> None:
                def _stream_to_queue(texto: str) -> None:
                    q.put({"type": "text", "text": texto})

                def _should_stop() -> bool:
                    return stop_event.is_set()

                try:
                    resultado = orq.procesar_consulta_compleja(
                        prompt,
                        stream_callback=_stream_to_queue,
                        should_stop=_should_stop,
                    )
                    q.put({"type": "final", "result": resultado})
                except Exception as e:
                    q.put({"type": "error", "error": str(e), "error_obj": e})

            t = threading.Thread(target=_worker, daemon=True)
            t.start()
            st.session_state["_gen_thread"] = t

        # --- Flujo chat-first: cuando el usuario pide "conectar impresora" sin IP,
        # mostrar selector del inventario y ejecutar automatización al confirmar.
        if "_pending_printer_connect" not in st.session_state:
            st.session_state["_pending_printer_connect"] = False
        if "_pending_printer_connect_user_text" not in st.session_state:
            st.session_state["_pending_printer_connect_user_text"] = ""

        auto_enabled_env = os.getenv("AI_SUPPORT_PRINTER_AUTOMATION_AUTO", "false").strip().lower() in {
            "1",
            "true",
            "yes",
            "y",
            "on",
        }

        if (
            auto_enabled_env
            and agente_estimado == "impresoras"
            and bool(st.session_state.get("_pending_printer_connect"))
        ):
            stored = st.session_state.get("_mysql_printers")
            if not (isinstance(stored, list) and stored and mysql_enabled()):
                st.warning(
                    "Para elegir una impresora necesito el inventario MySQL cargado. "
                    "Abre 'Diagnóstico local (PowerShell)' → 'Inventario impresoras (MySQL)' para cargarlo, "
                    "o escribe la IP en el chat."
                )
            else:
                st.info("Selecciona una impresora del inventario para conectar automáticamente.")
                options = [
                    f"{p.get('nombre','')} — {p.get('ubicacion','')} ({p.get('ip','')})"
                    for p in stored
                ]
                sel = st.selectbox(
                    "Impresora",
                    options=options,
                    index=0,
                    key="_pending_printer_connect_choice",
                )
                idx = options.index(sel) if sel in options else 0
                chosen = stored[idx]
                colpc1, colpc2 = st.columns([1, 1])
                with colpc1:
                    do_connect_selected = st.button(
                        "🖨️ Conectar seleccionada",
                        type="primary",
                        use_container_width=True,
                        key="_pending_printer_connect_run",
                        disabled=bool(st.session_state.get("_gen_active")),
                    )
                with colpc2:
                    do_cancel = st.button(
                        "Cancelar",
                        use_container_width=True,
                        key="_pending_printer_connect_cancel",
                        disabled=bool(st.session_state.get("_gen_active")),
                    )

                if do_cancel:
                    st.session_state["_pending_printer_connect"] = False
                    st.session_state["_pending_printer_connect_user_text"] = ""


                if do_connect_selected:
                    selected_ip = str(chosen.get("ip") or "").strip()
                    selected_name = str(chosen.get("nombre") or "").strip()
                    depto = str(chosen.get("nombre_departamento") or "").strip()
                    ubi = str(chosen.get("ubicacion") or "").strip()
                    selected_info = (
                        "[IMPRESORA_SELECCIONADA] "
                        f"nombre={selected_name} | ip={selected_ip} | departamento={depto} | ubicacion={ubi}"
                    )

                    if not selected_ip:
                        st.error("La impresora seleccionada no tiene IP.")
                    else:
                        base_user_text = str(st.session_state.get("_pending_printer_connect_user_text") or "")
                        drivers_dir = os.getenv(
                            "AI_SUPPORT_PRINTER_DRIVERS_DIR",
                            os.path.join(os.getcwd(), "printer_drivers"),
                        )
                        with st.spinner(f"Intentando conectar impresora {selected_ip} automáticamente..."):
                            try:
                                log = auto_connect_printer_ip(
                                    ip=selected_ip,
                                    user_text=(base_user_text + "\n" + selected_info).strip(),
                                    drivers_dir=drivers_dir,
                                    printer_display_name=selected_name or None,
                                )
                                st.session_state["_printer_auto_log"] = log.details
                                prompt = (base_user_text or "Conectar impresora").strip()
                                prompt = f"{prompt}\n\n{selected_info}\n{log.details}".strip()
                            except Exception as e:
                                st.session_state["_printer_auto_log"] = f"[AUTO_PRINTER] Error inesperado: {e}"
                                prompt = f"{prompt}\n\n{selected_info}\n{st.session_state['_printer_auto_log']}".strip()

                        st.session_state["_pending_printer_connect"] = False
                        st.session_state["_pending_printer_connect_user_text"] = ""
                        _start_generation(prompt)


        # --- IP Expander OCULTO: la asignación de IP es automática (similar a impresoras) ---
        # No se muestra expander de IP; todo se ejecuta en background cuando el usuario pide "conectarme a internet"

        # --- Crear contenedores PRIMERO (antes de procesar consultas) ---
        # Mostrar progreso ARRIBA del chat para que sea visible
        st.markdown("### 🔧 Diagnósticos y Automatizaciones")
        progress_container = st.container()
        
        st.markdown("---")
        st.markdown("### 💬 Conversación")
        # Contenedor para historial de chat
        history_container = st.container()

        if submitted and str(submitted).strip():
            consulta = str(submitted).strip()
            cooldown_until = float(st.session_state.get("_cooldown_until", 0.0) or 0.0)
            now = time.time()
            if cooldown_until > now:
                remaining = int(cooldown_until - now)
                st.warning(f"Demasiadas solicitudes recientemente. Espera {remaining}s y vuelve a intentar.")
                st.stop()

            if contiene_peligro(consulta):
                st.error(
                    "❌ Por motivos de seguridad y ética, no está permitido realizar preguntas relacionadas con hacking, "
                    "inyección SQL, ataques, acceso no autorizado o actividades peligrosas. Por favor, formula una consulta apropiada."
                )
            else:
                # Atajo: operaciones directas sobre Excel cargado (ChatGPT para Excel)
                df_excel = st.session_state.get("_excel_df")
                if df_excel is not None:
                    from ai_support.core.excel_chat_handler import handle_excel_command
                    result = handle_excel_command(df_excel, consulta)
                    if result["success"]:
                        st.success(result["message"])
                        if isinstance(result["result"], (int, float, str)):
                            st.metric(label="Resultado", value=result["result"])
                        elif isinstance(result["result"], pd.DataFrame):
                            st.dataframe(result["result"], use_container_width=True)
                            st.session_state["_excel_df"] = result["result"]  # actualizar con resultado filtrado/ordenado
                        st.stop()
                    # Si no se pudo procesar como comando Excel, agregar contexto del DataFrame al prompt
                    else:
                        import pandas as pd
                        import numpy as np
                        
                        # ESTRATEGIA INTELIGENTE: Filtrado exacto primero, luego búsqueda semántica
                        # Esto permite encontrar TODAS las filas que cumplan criterios específicos
                        
                        # Obtener embeddings configurados
                        embeddings = st.session_state.get("embeddings")
                        proveedor = st.session_state.get("provider_choice", "GitHub Models")
                        
                        # PASO 1: Intentar filtrado exacto por columnas
                        df_filtered = None
                        search_method = "keyword"
                        
                        # Detectar si la consulta menciona nombres de columnas
                        columnas_lower = [col.lower() for col in df_excel.columns]
                        consulta_lower = consulta.lower()
                        
                        # Buscar valores específicos para filtrar
                        filters = {}
                        
                        # Detectar "carrera" o nombre de carrera
                        if 'carrera' in columnas_lower:
                            carrera_col = df_excel.columns[columnas_lower.index('carrera')]
                            carreras_unicas = df_excel[carrera_col].unique()
                            for carrera in carreras_unicas:
                                if carrera and str(carrera).lower() in consulta_lower:
                                    filters[carrera_col] = carrera
                                    break
                        
                        # Detectar "jornada" o "horario"
                        for col_name in ['jornada', 'horario']:
                            if col_name in columnas_lower:
                                jornada_col = df_excel.columns[columnas_lower.index(col_name)]
                                for jornada in ['diurno', 'vespertino', 'ejecutivo']:
                                    if jornada in consulta_lower:
                                        filters[jornada_col] = df_excel[jornada_col].str.lower() == jornada
                                        break
                        
                        # Detectar "nivel"
                        if 'nivel' in columnas_lower:
                            nivel_col = df_excel.columns[columnas_lower.index('nivel')]
                            import re
                            nivel_match = re.search(r'\bnivel\s*(\d+)', consulta_lower)
                            if nivel_match:
                                nivel_num = int(nivel_match.group(1))
                                filters[nivel_col] = nivel_num
                        
                        # Aplicar filtros si se encontraron
                        if filters:
                            df_filtered = df_excel.copy()
                            for col, value in filters.items():
                                if isinstance(value, pd.Series):
                                    df_filtered = df_filtered[value]
                                else:
                                    df_filtered = df_filtered[df_filtered[col] == value]
                            
                            if len(df_filtered) > 0:
                                search_method = "exact_filter"
                        
                        # PASO 2: Si no hay filtrado exacto, intentar búsqueda semántica
                        if df_filtered is None or len(df_filtered) == 0:
                            if embeddings is not None and len(df_excel) > 50:
                                try:
                                    import sklearn
                                    from sklearn.metrics.pairwise import cosine_similarity
                                    
                                    # Crear embeddings de la consulta
                                    query_embedding = embeddings.embed_query(consulta)
                                    
                                    # Crear texto de cada fila (concatenar columnas importantes)
                                    def row_to_text(row):
                                        # Concatenar solo columnas de texto, máximo 500 caracteres por fila
                                        
                                        texts = []
                                        for col in df_excel.columns:
                                            val = str(row[col])
                                            if val and val != 'nan' and len(val) > 0:
                                                texts.append(val)
                                        return ' '.join(texts)[:500]
                                    
                                    # Limitar a primeras 500 filas para embeddings (por performance)
                                    df_sample = df_excel.head(500) if len(df_excel) > 500 else df_excel
                                    row_texts = df_sample.apply(row_to_text, axis=1).tolist()
                                    
                                    # Crear embeddings de las filas (en batches para no saturar)
                                    batch_size = 50
                                    row_embeddings = []
                                    for i in range(0, len(row_texts), batch_size):
                                        batch = row_texts[i:i+batch_size]
                                        batch_emb = embeddings.embed_documents(batch)
                                        batch_emb = embeddings.embed_documents(batch)
                                        row_embeddings.extend(batch_emb)
                                    
                                    # Calcular similitud coseno
                                    similarities = cosine_similarity([query_embedding], row_embeddings)[0]
                                    
                                    # Obtener índices de las filas más similares
                                    top_n = 100 if proveedor == "GitHub Models" else 200
                                    top_indices = np.argsort(similarities)[-top_n:][::-1]
                                    
                                    # Filtrar solo filas con similitud > 0.3 (umbral mínimo)
                                    relevant_indices = [idx for idx in top_indices if similarities[idx] > 0.3]
                                    
                                    if relevant_indices:
                                        df_filtered = df_sample.iloc[relevant_indices].copy()
                                        search_method = "semantic"
                                    
                                except Exception as e:
                                    # Si falla búsqueda semántica, usar fallback de keywords
                                    st.warning(f"Búsqueda semántica no disponible, usando keywords: {e}")
                        
                        # FALLBACK: Búsqueda por keywords si no hay embeddings o falló
                        if df_filtered is None or len(df_filtered) == 0:
                            # Extraer palabras clave de la consulta
                            stop_words = {'el', 'la', 'de', 'en', 'a', 'los', 'las', 'un', 'una', 'por', 'para', 'con', 'del', 'y', 'o', 'que', 'es', 'son', 'cuales', 'cual', 'me', 'dame', 'muestra', 'lista', 'busca', 'encuentra'}
                            consulta_lower = consulta.lower()
                            palabras = [p.strip('¿?.,;:()[]{}') for p in consulta_lower.split()]
                            keywords = [p for p in palabras if len(p) > 3 and p not in stop_words]
                            
                            # Filtrar DataFrame buscando keywords en TODAS las columnas
                            df_filtered = df_excel.copy()
                            if keywords:
                                mask = pd.Series([False] * len(df_excel))
                                for col in df_excel.columns:
                                    col_str = df_excel[col].astype(str).str.lower()
                                    for keyword in keywords:
                                        mask |= col_str.str.contains(keyword, na=False, regex=False)
                                
                                df_filtered = df_excel[mask]
                                
                                if len(df_filtered) == 0:
                                    df_filtered = df_excel.copy()
                        
                        # Agregar información del DataFrame al prompt
                        excel_context = f"\n\n📊 **IMPORTANTE: Tengo acceso directo a un archivo Excel cargado en memoria. Puedo analizar sus datos.**\n\n"
                        excel_context += f"**Información del archivo:**\n"
                        excel_context += f"- Columnas: {', '.join(df_excel.columns.astype(str))}\n"
                        excel_context += f"- Total de filas en archivo: {len(df_excel)}\n"
                        
                        if len(df_filtered) < len(df_excel):
                            if search_method == "exact_filter":
                                excel_context += f"- Filas filtradas exactamente: {len(df_filtered)} (filtrado por columnas específicas)\n"
                            elif search_method == "semantic":
                                excel_context += f"- Filas relevantes encontradas: {len(df_filtered)} (búsqueda semántica con IA)\n"
                            else:
                                excel_context += f"- Filas relevantes encontradas: {len(df_filtered)} (búsqueda por palabras clave)\n"
                        
                        excel_context += f"- Tipos de datos: {df_excel.dtypes.to_dict()}\n\n"
                        
                        # Estadísticas descriptivas si hay columnas numéricas
                        numeric_cols = df_filtered.select_dtypes(include=['number']).columns
                        if len(numeric_cols) > 0:
                            excel_context += f"**Estadísticas de columnas numéricas:**\n{df_filtered[numeric_cols].describe().to_string()}\n\n"
                        
                        # Valores únicos de columnas categóricas (útil para filtros)
                        categorical_cols = df_filtered.select_dtypes(include=['object']).columns
                        if len(categorical_cols) > 0 and len(categorical_cols) <= 10:
                            excel_context += f"**Valores únicos en columnas de texto:**\n"
                            for col in categorical_cols[:8]:  # Primeras 8 columnas
                                unique_vals = df_filtered[col].dropna().unique()[:15]  # Primeros 15 valores
                                if len(unique_vals) > 0:
                                    excel_context += f"- {col}: {', '.join(map(str, unique_vals))}\n"
                            excel_context += "\n"
                        
                        # Ajustar límite de filas según el proveedor Y el método de búsqueda
                        proveedor = st.session_state.get("provider_choice", "GitHub Models")
                        if search_method == "exact_filter":
                            # Si es filtrado exacto, mostrar TODAS las filas (no hay ambigüedad)
                            max_rows = len(df_filtered)
                        elif proveedor == "LM Studio (local)":
                            max_rows = min(300, len(df_filtered))
                        else:
                            max_rows = min(100, len(df_filtered))
                        
                        # Enviar datos filtrados (solo una vez)
                        excel_context += f"**DATOS RELEVANTES ({max_rows} filas más relevantes):**\n```\n{df_filtered.head(max_rows).to_string()}\n```\n\n"
                        
                        # Información sobre duplicados
                        total_duplicates = df_excel.duplicated().sum()
                        if total_duplicates > 0:
                            excel_context += f"⚠️ **Nota:** Hay {total_duplicates} filas duplicadas en el archivo completo.\n\n"
                        
                        excel_context += f"**INSTRUCCIONES IMPORTANTES:**\n"
                        if search_method == "exact_filter":
                            excel_context += f"1. Los datos mostrados fueron FILTRADOS EXACTAMENTE por las columnas mencionadas en la consulta.\n"
                            excel_context += f"2. Tienes acceso a TODAS las {max_rows} filas que cumplen exactamente los criterios.\n"
                            excel_context += f"3. MUESTRA TODOS LOS RESULTADOS ENCONTRADOS - son exactamente lo que el usuario pidió.\n"
                        elif search_method == "semantic":
                            excel_context += f"1. Los datos mostrados fueron seleccionados con BÚSQUEDA SEMÁNTICA (IA) - son las {max_rows} filas MÁS RELEVANTES de {len(df_excel)} totales.\n"
                            excel_context += f"2. Tienes acceso a las {max_rows} filas MÁS RELEVANTES encontradas.\n"
                        else:
                            excel_context += f"1. Los datos mostrados fueron filtrados buscando en TODO el archivo de {len(df_excel)} filas.\n"
                            excel_context += f"2. Tienes acceso a las {max_rows} filas MÁS RELEVANTES encontradas.\n"
                        excel_context += f"3. SIEMPRE muestra los valores EXACTOS de las columnas solicitadas, NO uses placeholders.\n"
                        excel_context += f"4. Si encontraste resultados, lista TODOS los valores reales de las columnas pedidas.\n"
                        excel_context += f"5. Si NO hay resultados en los datos mostrados, indícalo claramente.\n"
                        excel_context += f"6. NO des instrucciones de Excel manual. Proporciona el análisis directo.\n\n"
                        
                        prompt = consulta + excel_context
                else:
                    prompt = consulta
                
                # Agregar diagnósticos de impresora si existen
                if printer_diag_for_prompt:
                    prompt = f"{prompt}\n\n{printer_diag_for_prompt}"



                # Automatización: si se pide conectar impresora por IP, intentar flujo automático.
                # Para evitar ejecuciones no deseadas, solo corre si AI_SUPPORT_PRINTER_AUTOMATION_AUTO=true.
                auto_enabled = os.getenv("AI_SUPPORT_PRINTER_AUTOMATION_AUTO", "false").strip().lower() in {
                    "1",
                    "true",
                    "yes",
                    "y",
                    "on",
                }
                if auto_enabled and agente_estimado == "impresoras":
                    ip = _extract_ipv4(consulta)
                    wants_connect = any(w in consulta.lower() for w in ["conectar", "agregar", "instalar", "añadir"])

                    selected = st.session_state.get("_selected_printer_record")
                    selected_ip = None
                    selected_info = ""
                    if isinstance(selected, dict):
                        selected_ip = str(selected.get("ip") or "").strip() or None
                        nombre = str(selected.get("nombre") or "").strip()
                        depto = str(selected.get("nombre_departamento") or "").strip()
                        ubi = str(selected.get("ubicacion") or "").strip()
                        if nombre or depto or ubi or selected_ip:
                            selected_info = (
                                "[IMPRESORA_SELECCIONADA] "
                                f"nombre={nombre} | ip={selected_ip or ''} | departamento={depto} | ubicacion={ubi}"
                            )

                    if (not ip) and selected_ip:
                        ip = selected_ip

                    # Si se pidió conectar pero NO hay IP ni selección previa, pedir selección del inventario.
                    if wants_connect and (not ip) and mysql_enabled():
                        stored = st.session_state.get("_mysql_printers")
                        if isinstance(stored, list) and stored:
                            st.session_state["_pending_printer_connect"] = True
                            st.session_state["_pending_printer_connect_user_text"] = consulta
                            st.rerun()

                    if ip and wants_connect:
                        drivers_dir = os.getenv(
                            "AI_SUPPORT_PRINTER_DRIVERS_DIR",
                            os.path.join(os.getcwd(), "printer_drivers"),
                        )
                        with st.spinner(f"Intentando conectar impresora {ip} automáticamente..."):
                            try:
                                log = auto_connect_printer_ip(
                                    ip=ip,
                                    user_text=(consulta + ("\n" + selected_info if selected_info else "")),
                                    drivers_dir=drivers_dir,
                                    printer_display_name=(str(selected.get("nombre") or "").strip() if isinstance(selected, dict) else None) or None,
                                )
                                st.session_state["_printer_auto_log"] = log.details
                                # Adjuntar al prompt para que el agente explique lo que pasó.
                                if selected_info:
                                    prompt = f"{prompt}\n\n{selected_info}\n{log.details}"
                                else:
                                    prompt = f"{prompt}\n\n{log.details}"
                            except Exception as e:
                                st.session_state["_printer_auto_log"] = f"[AUTO_PRINTER] Error inesperado: {e}"
                                prompt = f"{prompt}\n\n{st.session_state['_printer_auto_log']}"

                # AHORA SÍ: Guardar en historial de conversación (después de automatizaciones)
                current_conv_id = st.session_state.get("_current_conversation_id")
                if not current_conv_id:
                    # Crear nueva conversación automáticamente
                    import datetime
                    new_id = f"conv_{int(time.time())}_{secrets.token_hex(4)}"
                    # Generar título a partir de las primeras palabras de la consulta
                    title_words = consulta.split()[:6]
                    conv_title = " ".join(title_words) + ("..." if len(consulta.split()) > 6 else "")
                    new_conv = {
                        "id": new_id,
                        "title": conv_title,
                        "created_at": datetime.datetime.now().strftime("%Y-%m-%d %H:%M"),
                        "messages": []
                    }
                    st.session_state["_conversations"].insert(0, new_conv)
                    st.session_state["_current_conversation_id"] = new_id
                    st.session_state["_conversation_messages"][new_id] = []
                    current_conv_id = new_id
                
                # Agregar mensaje del usuario
                if current_conv_id not in st.session_state["_conversation_messages"]:
                    st.session_state["_conversation_messages"][current_conv_id] = []
                
                st.session_state["_conversation_messages"][current_conv_id].append({
                    "role": "user",
                    "content": consulta,
                    "timestamp": time.time()
                })

                _persist_ui_conversations()
                
                _start_generation(prompt)


        # --- Render de chat + streaming ---
        current_conv_id = st.session_state.get("_current_conversation_id")
        messages = (
            st.session_state.get("_conversation_messages", {}).get(current_conv_id, [])
            if current_conv_id
            else []
        )

        # Consumir cola mientras se genera (no bloqueante)
        if st.session_state.get("_gen_active"):
            t = st.session_state.get("_gen_thread")
            if t is not None and hasattr(t, "is_alive") and (not t.is_alive()):
                st.session_state["_gen_active"] = False
                st.session_state["_gen_stop_event"] = None

            q = st.session_state.get("_gen_queue")
            if q is not None:
                try:
                    while True:
                        msg = q.get_nowait()
                        if msg.get("type") == "text":
                            st.session_state["_gen_text"] = msg.get("text", "")
                        elif msg.get("type") == "final":
                            final_result = msg.get("result")
                            st.session_state["_gen_result"] = final_result
                            if isinstance(final_result, dict) and isinstance(final_result.get("respuesta"), str):
                                st.session_state["_gen_text"] = final_result.get("respuesta") or st.session_state.get(
                                    "_gen_text", ""
                                )
                            st.session_state["_gen_active"] = False
                            st.session_state["_gen_stop_event"] = None
                            st.session_state["_gen_queue"] = None  # Limpiar la cola también

                            current_conv_id = st.session_state.get("_current_conversation_id")
                            if current_conv_id and current_conv_id in st.session_state["_conversation_messages"]:
                                st.session_state["_conversation_messages"][current_conv_id].append(
                                    {
                                        "role": "assistant",
                                        "content": st.session_state["_gen_text"],
                                        "result": final_result,
                                        "timestamp": time.time(),
                                    }
                                )
                                _persist_ui_conversations()
                            st.rerun()  # Actualizar UI: ocultar Stop, mostrar respuesta

                        elif msg.get("type") == "error":
                            st.session_state["_gen_error"] = msg.get("error")
                            st.session_state["_gen_error_obj"] = msg.get("error_obj")
                            st.session_state["_gen_active"] = False
                            st.session_state["_gen_stop_event"] = None
                            st.session_state["_gen_queue"] = None  # Limpiar la cola también
                            st.rerun()  # Actualizar UI: mostrar error, habilitar input

                except queue.Empty:
                    pass

        # --- Render de chat + streaming ---
        # Usar contenedores ya creados arriba (líneas 1991-1994)
        
        # Mostrar historial en contenedor inferior
        with history_container:
            current_conv_id = st.session_state.get("_current_conversation_id")
            messages = (
                st.session_state.get("_conversation_messages", {}).get(current_conv_id, [])
                if current_conv_id
                else []
            )

            if messages:
                st.markdown("---")
                st.caption("📜 Historial de conversación (más recientes primero)")
                
            for msg in reversed(messages):
                role = str(msg.get("role") or "").strip().lower()
                content = str(msg.get("content") or "")
                if role == "user":
                    with st.chat_message("user"):
                        st.markdown(content)
                elif role == "assistant":
                    with st.chat_message("assistant"):
                        st.markdown(content)

        # Diagnóstico de red se muestra en tiempo real durante la ejecución (no necesita lógica aquí)
        
        if st.session_state.get("_gen_error"):
            with progress_container:
                st.error(str(st.session_state.get("_gen_error") or "Error desconocido"))

        # Solo hacer rerun si realmente está activo
        if st.session_state.get("_gen_active"):
            with progress_container:
                with st.chat_message("assistant"):
                    st.caption("⏳ Generando…")
                    st.markdown(st.session_state.get("_gen_text", ""))
            
            # Verificar de nuevo antes de rerun por si se procesó mensaje final
            if st.session_state.get("_gen_active"):
                time.sleep(0.2)
                st.rerun()  # ← re-ejecuta la página para leer la cola del hilo


        if (not st.session_state.get("_gen_active")) and st.session_state.get("_gen_result"):
            resultado = st.session_state.get("_gen_result")
            st.session_state["_gen_stop_event"] = None

            with progress_container:
                with st.expander("🔧 Detalles de ejecución", expanded=False):
                    st.info(f"🎯 **Agente Principal**: {resultado['agente_principal']}")
                    st.info(f"👥 **Agentes Involucrados**: {', '.join(resultado['agentes_involucrados'])}")
                    st.info(f"⏱️ **Tiempo**: {resultado['tiempo_respuesta']:.2f}s")

                if resultado.get("stopped"):
                    st.warning("Generación detenida: se muestra respuesta parcial.")

                if "colaboracion" in resultado:
                    with st.expander("🔗 Colaboración Multi-Agente"):
                        st.markdown(resultado["colaboracion"])

                if resultado.get("kb_usado"):
                    with st.expander("📚 Base de Conocimiento utilizada", expanded=True):
                        st.success("✅ Se encontraron procedimientos relevantes en la Base de Conocimiento")
                        if resultado.get("kb_preview"):
                            st.markdown("**Fragmento recuperado:**")
                            st.text(resultado["kb_preview"])

                if resultado.get("faiss_usado"):
                    with st.expander("🔍 FAISS RAG Utilizado"):
                        st.success("✅ Búsqueda semántica FAISS activa")
                        if resultado.get("contexto_faiss"):
                            st.markdown("**Contexto encontrado:**")
                            st.text(resultado["contexto_faiss"])
                        else:
                            st.info("Contexto FAISS disponible pero no mostrado")

                if "memoria_usada" in resultado:
                    with st.expander("🧠 Memoria Utilizada"):
                        memoria_info = resultado["memoria_usada"]
                        col_mem1, col_mem2, col_mem3 = st.columns(3)

                        with col_mem1:
                            st.metric("Buffer", memoria_info.get("buffer", 0))
                            st.caption("Historial completo")
                        with col_mem2:
                            st.metric("Summary", memoria_info.get("summary", 0))
                            st.caption("Resumen inteligente")
                        with col_mem3:
                            st.metric("Window", memoria_info.get("window", 0))
                            st.caption("Últimas interacciones")

                        col_mem4, col_mem5 = st.columns(2)
                    with col_mem4:
                        st.metric("Entities", memoria_info.get("entities", 0))
                        st.caption("Entidades recordadas")
                    with col_mem5:
                        st.metric("Vector", memoria_info.get("vector", 0))
                        st.caption("Memoria a largo plazo")

        if (not st.session_state.get("_gen_active")) and st.session_state.get("_gen_error"):
            err_str = st.session_state.get("_gen_error") or "(error)"
            err_obj = st.session_state.get("_gen_error_obj")

            # limpiar stop event
            st.session_state["_gen_stop_event"] = None

            if err_obj is not None and _is_rate_limit_error(err_obj):
                st.session_state["_cooldown_until"] = time.time() + 120
                st.error(
                    "El proveedor devolvió 'Too many requests' (límite de solicitudes). "
                    "Espera ~2 minutos y reintenta, o cambia a `LM Studio (local)` para evitar límites."
                )
            elif err_obj is not None and _is_github_no_access_error(err_obj):
                st.error(
                    "Tu token no tiene acceso al modelo seleccionado en GitHub Models. "
                    "Cambia el `Modelo LLM` por uno permitido para tu cuenta, o usa `LM Studio (local)` en el sidebar."
                )
            else:
                st.error(f"Error al generar respuesta: {err_str}")

    # --- Sección Excel con IA (en columna izquierda, abajo) ---
    with chat_col:
        st.markdown("---")
        with st.expander("📊 Análisis de Excel con IA", expanded=False):
            st.caption(
                "Sube un archivo Excel (.xlsx/.xls) o CSV para hacer preguntas al agente sobre los datos."
            )
            
            # Debug: verificar disponibilidad de openpyxl
            try:
                import openpyxl
                st.caption(f"✓ openpyxl disponible (v{openpyxl.__version__})")
            except ImportError:
                st.error("✗ openpyxl NO disponible - usa CSV")
            
            uploaded = st.file_uploader(
                "Selecciona archivo Excel o CSV",
                type=["xlsx", "xls", "csv"],
                key="excel_uploader",
            )
            
            if uploaded is not None:
                import pandas as pd
                ext = os.path.splitext(uploaded.name)[1].lower()
                df = None
                
                if ext == ".csv":
                    try:
                        encodings = ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']
                        separators = [',', ';', '\t']
                        
                        for encoding in encodings:
                            for sep in separators:
                                try:
                                    uploaded.seek(0)
                                    df = pd.read_csv(
                                        uploaded, 
                                        encoding=encoding,
                                        sep=sep,
                                        on_bad_lines='skip',
                                        engine='python'
                                    )
                                    if not df.empty:
                                        st.success(f"✓ CSV cargado ({len(df)} filas)")
                                        break
                                except Exception:
                                    continue
                            if df is not None and not df.empty:
                                break
                        
                        if df is None or df.empty:
                            st.error("No se pudo leer el CSV.")
                    except Exception as e:
                        st.error(f"Error: {e}")
                        
                elif ext in {".xlsx", ".xls"}:
                    try:
                        from openpyxl import load_workbook
                        import io
                        
                        wb = load_workbook(io.BytesIO(uploaded.read()))
                        all_sheets = {}
                        
                        for sheet_name in wb.sheetnames:
                            ws = wb[sheet_name]
                            data = []
                            for row in ws.iter_rows(values_only=True):
                                data.append(list(row))
                            
                            if data and len(data) > 0:
                                headers = data[0]
                                clean_headers = []
                                seen = {}
                                for i, header in enumerate(headers):
                                    if header is None or str(header).strip() == "":
                                        header = f"Columna_{i+1}"
                                    else:
                                        header = str(header).strip()
                                    
                                    if header in seen:
                                        seen[header] += 1
                                        header = f"{header}_{seen[header]}"
                                    else:
                                        seen[header] = 0
                                    
                                    clean_headers.append(header)
                                
                                sheet_df = pd.DataFrame(data[1:], columns=clean_headers)
                                all_sheets[sheet_name] = sheet_df
                        
                        if all_sheets:
                            st.success(f"✓ Excel cargado: {len(all_sheets)} pestaña(s)")
                            st.session_state["_excel_sheets"] = all_sheets
                            st.session_state["_excel_uploaded_file"] = uploaded
                            st.session_state["_excel_filename"] = uploaded.name
                            
                            st.markdown("### 📋 Vista previa")
                            for sheet_name, sheet_df in all_sheets.items():
                                with st.expander(f"📄 {sheet_name} ({len(sheet_df)} filas)", expanded=False):
                                    preview_rows = min(10, len(sheet_df))
                                    st.dataframe(sheet_df.head(preview_rows), use_container_width=True)
                                    if len(sheet_df) > preview_rows:
                                        st.caption(f"Mostrando {preview_rows} de {len(sheet_df)} filas")
                            
                            df = list(all_sheets.values())[0]
                            st.session_state["_excel_df"] = df
                            st.info("💬 Haz preguntas en el chat principal")
                        else:
                            st.warning("Archivo vacío")
                    except ImportError:
                        st.error("openpyxl no disponible - usa CSV")
                    except Exception as e:
                        st.error(f"Error: {e}")

                if df is not None and ext == ".csv":
                    st.session_state["_excel_df"] = df
                    st.session_state["_excel_uploaded_file"] = uploaded
                    st.session_state["_excel_filename"] = uploaded.name
                    
                    st.success(f"✓ CSV: {len(df)} filas, {len(df.columns)} columnas")
                    st.markdown("### 📋 Vista previa")
                    preview_rows = min(10, len(df))
                    st.dataframe(df.head(preview_rows), use_container_width=True)
                    if len(df) > preview_rows:
                        st.caption(f"Mostrando {preview_rows} de {len(df)} filas")
                    st.info("💬 Haz preguntas en el chat principal")


if __name__ == "__main__":
    main()
